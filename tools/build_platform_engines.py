#!/usr/bin/env python3
"""Engine workbooks for the cross-section platforms (Athene-parity artifacts).

One stock-and-flow workbook per platform, same discipline as
athene-quarterly-engine.xlsx / ga-quarterly-engine.xlsx: blue = value as
printed in the source filing, black/bold = live formula, Checks tab = master
count of failing identities (must read 0). Every identity is also verified in
Python at build time (no LibreOffice on this machine — openpyxl writes
formulas without cached values, so the sheet-side checks are for Excel users).

  brookfield -> dossiers/brookfield/bnt-quarterly-engine.xlsx  (Engine 4Q24-1Q26 + Engine-FY 2023-25)
  ares       -> dossiers/ares/aspida-annual-engine.xlsx        (Engine-FY 2023-25; no quarterly cadence exists)
  blueowl    -> dossiers/blueowl/klr-annual-engine.xlsx        (Engine-FY 2022-25)

Usage: build_platform_engines.py [bnt|aspida|klr ...]  (default: all)
"""
import csv
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10)
BOLD = Font(name='Arial', size=10, bold=True)
HDR = Font(name='Arial', size=11, bold=True)
GREY = Font(name='Arial', size=9, color='666666')
MONEY = '#,##0;(#,##0);"-"'
PCT = '0.00%'


def load(fname, pref):
    """(period, metric) -> value, preferring sources by substring order in pref."""
    by = defaultdict(dict)
    for r in csv.DictReader(open(ROOT / f'extract/cross-section/{fname}.csv')):
        by[(r['period'], r['metric'])][r['source_file']] = float(r['value'])
    out = {}
    for k, srcs in by.items():
        pick = sorted(srcs, key=lambda s: next((i for i, p in enumerate(pref) if p in s), 99))[0]
        out[k] = srcs[pick]
    return out


class Builder:
    def __init__(self, wb, sheet, periods, data, tol=2):
        self.ws = wb.create_sheet(sheet)
        self.periods = periods
        self.D = data
        self.tol = tol
        self.row = 0
        self.checks = []          # (sheet, cell) of every check formula
        self.py_fail = []
        self.marks = {}           # label -> row (for formulas referencing earlier rows)
        self.ws.column_dimensions['A'].width = 58
        for i in range(len(periods)):
            self.ws.column_dimensions[get_column_letter(2 + i)].width = 14
        self._head()

    def _head(self):
        self.row += 1
        self.ws.cell(row=self.row, column=1, value='').font = BLACK
        for i, p in enumerate(self.periods):
            c = self.ws.cell(row=self.row, column=2 + i, value=p)
            c.font = HDR

    def section(self, title):
        self.row += 2
        self.ws.cell(row=self.row, column=1, value=title).font = HDR

    def note(self, text):
        self.row += 1
        self.ws.cell(row=self.row, column=1, value=text).font = GREY

    def vals(self, label, metric, fmt=MONEY, indent=0, mark=None, scale=1.0):
        """Blue printed-value row; missing period -> blank cell (boundary)."""
        self.row += 1
        self.ws.cell(row=self.row, column=1, value='    ' * indent + label).font = BLACK
        for i, p in enumerate(self.periods):
            v = self.D.get((p, metric))
            c = self.ws.cell(row=self.row, column=2 + i)
            if v is not None:
                c.value = v * scale
            c.font = BLUE
            c.number_format = fmt
        self.marks[mark or label] = self.row
        return self.row

    def formula(self, label, fn, fmt=MONEY, mark=None, font=BOLD):
        self.row += 1
        self.ws.cell(row=self.row, column=1, value=label).font = font
        for i in range(len(self.periods)):
            col = get_column_letter(2 + i)
            c = self.ws.cell(row=self.row, column=2 + i, value=fn(col))
            c.font = font
            c.number_format = fmt
        self.marks[mark or label] = self.row
        return self.row

    def check(self, label, lhs_fn, rhs_fn, py_lhs=None, py_rhs=None, tol=None):
        """CHECK row: OK/FAIL per period; n/a when either side is blank.
        py_lhs/py_rhs: per-period python values to verify the identity at build time."""
        tol = self.tol if tol is None else tol
        self.row += 1
        self.ws.cell(row=self.row, column=1, value='CHECK: ' + label).font = GREY
        for i, p in enumerate(self.periods):
            col = get_column_letter(2 + i)
            l, r = lhs_fn(col), rhs_fn(col)
            f = (f'=IF(OR({l}="",{r}=""),"n/a",'
                 f'IF(ABS({l}-{r})<={tol},"OK","FAIL"))')
            c = self.ws.cell(row=self.row, column=2 + i, value=f)
            c.font = GREY
            self.checks.append((self.ws.title, f'{col}{self.row}'))
            if py_lhs is not None and py_rhs is not None:
                a, b = py_lhs(p), py_rhs(p)
                if a is not None and b is not None and abs(a - b) > tol:
                    self.py_fail.append(f'{self.ws.title} {label} {p}: {a:,.0f} vs {b:,.0f}')

    def ref(self, mark, col):
        return f'{col}{self.marks[mark]}'

    def sumrows(self, marks, col):
        return '+'.join(f'{col}{self.marks[m]}' for m in marks)


def checks_sheet(wb, all_checks, title_note):
    ws = wb.create_sheet('Checks')
    ws.column_dimensions['A'].width = 70
    ws.cell(row=1, column=1, value='Master check — count of FAIL cells across all engine tabs (must be 0)').font = HDR
    refs = [f"'{s}'!{c}" for s, c in all_checks]
    parts, chunk = [], []
    for r in refs:
        chunk.append(r)
        if sum(len(x) + 22 for x in chunk) > 1000:
            parts.append(chunk[:-1])
            chunk = [r]
    parts.append(chunk)
    prow = 3
    prows = []
    for part in parts:
        f = '=' + '+'.join(f'COUNTIF({r},"FAIL")' for r in part)
        ws.cell(row=prow, column=1, value=f'chunk {len(prows) + 1}').font = GREY
        ws.cell(row=prow, column=2, value=f).font = BLACK
        prows.append(prow)
        prow += 1
    ws.cell(row=2, column=1, value='TOTAL FAILING CHECKS').font = BOLD
    ws.cell(row=2, column=2, value='=' + '+'.join(f'B{r}' for r in prows)).font = BOLD
    ws.cell(row=prow + 1, column=1, value=f'{len(refs)} check cells tracked. {title_note}').font = GREY


def readme(wb, lines):
    ws = wb.create_sheet('ReadMe', 0)
    ws.column_dimensions['A'].width = 120
    for i, ln in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=ln)
        c.font = HDR if i == 1 else BLACK
        if ln.startswith('='):
            c.data_type = 's'


# ---------------------------------------------------------------- Brookfield
def build_bnt():
    D = load('bnt_engine_raw', ['20f', 'q1-2026', 'q4-2025'])
    wb = Workbook()
    wb.remove(wb.active)
    g = lambda p, m: D.get((p, m))

    # ---- Engine-FY ----
    b = Builder(wb, 'Engine-FY', ['FY2023', 'FY2024', 'FY2025'], D)
    b.section('① MONEY IN — annuity sales and net premiums (US$ millions)')
    b.vals('Retail annuity sales — fixed index', 'gross_annuity_sales_fixed_index', indent=1)
    b.vals('Retail annuity sales — fixed rate', 'gross_annuity_sales_fixed_rate', indent=1)
    b.vals('Retail annuity sales — variable/other', 'gross_annuity_sales_variable', indent=1)
    b.vals('Retail annuity sales (printed)', 'gross_annuity_sales_retail', mark='retail')
    b.check('retail sales foot',
            lambda c: b.sumrows(['Retail annuity sales — fixed index', 'Retail annuity sales — fixed rate',
                                 'Retail annuity sales — variable/other'], c),
            lambda c: b.ref('retail', c),
            lambda p: sum(g(p, m) or 0 for m in ('gross_annuity_sales_fixed_index', 'gross_annuity_sales_fixed_rate',
                                                 'gross_annuity_sales_variable')),
            lambda p: g(p, 'gross_annuity_sales_retail'))
    b.vals('Institutional — PRT', 'gross_annuity_sales_prt', indent=1)
    b.vals('Institutional — funding agreements', 'gross_annuity_sales_funding_agreements', indent=1)
    b.vals('Institutional annuity sales (printed)', 'gross_annuity_sales_institutional', mark='inst')
    b.check('institutional sales foot',
            lambda c: b.sumrows(['Institutional — PRT', 'Institutional — funding agreements'], c),
            lambda c: b.ref('inst', c),
            lambda p: sum(g(p, m) or 0 for m in ('gross_annuity_sales_prt', 'gross_annuity_sales_funding_agreements')),
            lambda p: g(p, 'gross_annuity_sales_institutional'))
    b.vals('TOTAL GROSS ANNUITY SALES (printed)', 'gross_annuity_sales_total', mark='sales_tot')
    b.check('retail + institutional = total sales',
            lambda c: f"{b.ref('retail', c)}+{b.ref('inst', c)}",
            lambda c: b.ref('sales_tot', c),
            lambda p: (g(p, 'gross_annuity_sales_retail') or 0) + (g(p, 'gross_annuity_sales_institutional') or 0),
            lambda p: g(p, 'gross_annuity_sales_total'))
    b.vals('Net premiums — annuities', 'net_premiums_total_annuities', indent=1)
    b.vals('Net premiums — P&C', 'net_premiums_total_pc', indent=1)
    b.vals('Net premiums — whole life & others', 'net_premiums_whole_life_and_others', indent=1)
    b.vals('NET PREMIUMS (printed)', 'net_premiums', mark='netprem')
    b.check('net premiums by product foot',
            lambda c: b.sumrows(['Net premiums — annuities', 'Net premiums — P&C',
                                 'Net premiums — whole life & others'], c),
            lambda c: b.ref('netprem', c),
            lambda p: sum(g(p, m) or 0 for m in ('net_premiums_total_annuities', 'net_premiums_total_pc',
                                                 'net_premiums_whole_life_and_others')),
            lambda p: g(p, 'net_premiums'))
    b.vals("Policyholders' account deposits (cash flow)", 'policyholders_account_deposits')
    b.vals("Policyholders' account withdrawals (cash flow)", 'policyholders_account_withdrawals')

    b.section('② REVENUES (US GAAP)')
    b.vals('Net premiums', 'net_premiums', indent=1)
    b.vals('Other policy revenue', 'other_policy_revenue', indent=1)
    b.vals('Net investment income', 'net_investment_income', indent=1)
    b.vals('Investment related gains', 'investment_related_gains', indent=1)
    b.vals('Net investment results — reinsurance funds withheld', 'net_investment_results_reinsurance_funds_withheld', indent=1)
    b.vals('TOTAL REVENUES (printed)', 'total_revenues', mark='rev_tot')
    REV = ('net_premiums', 'other_policy_revenue', 'net_investment_income', 'investment_related_gains',
           'net_investment_results_reinsurance_funds_withheld')
    b.check('revenue lines foot',
            lambda c: b.sumrows(['Net premiums', 'Other policy revenue', 'Net investment income',
                                 'Investment related gains',
                                 'Net investment results — reinsurance funds withheld'], c),
            lambda c: b.ref('rev_tot', c),
            lambda p: sum(g(p, m) or 0 for m in REV), lambda p: g(p, 'total_revenues'))

    b.section('③ BENEFITS AND EXPENSES (stored signed as printed — parentheses negative)')
    EXP = ('policyholder_benefits_and_claims_incurred', 'interest_sensitive_contract_benefits',
           'amortization_dac_dsi_voba', 'operating_expenses', 'interest_expense', 'other_reinsurance_expenses',
           'change_fv_insurance_derivatives_embedded_derivatives', 'change_fv_market_risk_benefits')
    labels = ['Policyholder benefits and claims incurred', 'Interest sensitive contract benefits',
              'Amortization of DAC / DSI / VOBA', 'Operating expenses', 'Interest expense',
              'Other reinsurance expenses', 'Change in FV — insurance & embedded derivatives',
              'Change in FV — market risk benefits']
    for lab, m in zip(labels, EXP):
        b.vals(lab, m, indent=1)
    b.vals('TOTAL BENEFITS AND EXPENSES (printed)', 'total_benefits_and_expenses', mark='exp_tot')
    b.check('expense lines foot',
            lambda c: b.sumrows(labels, c), lambda c: b.ref('exp_tot', c),
            lambda p: sum(g(p, m) or 0 for m in EXP), lambda p: g(p, 'total_benefits_and_expenses'))

    b.section('④ NET INCOME')
    b.vals('Net income before income taxes (printed)', 'net_income_before_income_taxes', mark='pretax')
    b.check('revenues + expenses = pretax',
            lambda c: f"{b.ref('rev_tot', c)}+{b.ref('exp_tot', c)}",
            lambda c: b.ref('pretax', c),
            lambda p: (g(p, 'total_revenues') or 0) + (g(p, 'total_benefits_and_expenses') or 0),
            lambda p: g(p, 'net_income_before_income_taxes'))
    b.vals('Income tax recovery (expense)', 'income_tax_recovery_expense', mark='tax')
    b.vals('NET INCOME (printed)', 'net_income', mark='ni')
    b.check('pretax + tax = net income',
            lambda c: f"{b.ref('pretax', c)}+{b.ref('tax', c)}", lambda c: b.ref('ni', c),
            lambda p: (g(p, 'net_income_before_income_taxes') or 0) + (g(p, 'income_tax_recovery_expense') or 0),
            lambda p: g(p, 'net_income'))
    b.vals('  attributable to Class A/B shareholders', 'net_income_attrib_class_a_b_shareholders', indent=1)
    b.vals('  attributable to Class C shareholder', 'net_income_attrib_class_c_shareholder', indent=1)
    b.vals('  attributable to NCI', 'net_income_attrib_nci', indent=1)
    b.check('attribution foots',
            lambda c: b.sumrows(['  attributable to Class A/B shareholders', '  attributable to Class C shareholder',
                                 '  attributable to NCI'], c),
            lambda c: b.ref('ni', c),
            lambda p: sum(g(p, m) or 0 for m in ('net_income_attrib_class_a_b_shareholders',
                                                 'net_income_attrib_class_c_shareholder', 'net_income_attrib_nci')),
            lambda p: g(p, 'net_income'))
    b.vals('Net income attributable to shareholders (MD&A)', 'net_income_attributable_to_shareholders')

    b.section("⑤ DISTRIBUTABLE OPERATING EARNINGS — BNT's SRE analog (non-GAAP, printed reconciliation)")
    RECON = ('doe_recon_mtm_investments_incl_fw', 'doe_recon_mtm_insurance_contracts',
             'doe_recon_deferred_tax_basis_changes', 'doe_recon_depreciation_amortization',
             'doe_recon_transaction_costs')
    rl = ['  add back: MTM on investments incl funds withheld', '  add back: MTM on insurance contracts',
          '  add back: deferred tax / basis changes', '  add back: depreciation & amortization',
          '  add back: transaction costs']
    for lab, m in zip(rl, RECON):
        b.vals(lab, m, indent=1)
    b.vals('DISTRIBUTABLE OPERATING EARNINGS (printed)', 'distributable_operating_earnings', mark='doe')
    b.check('NI + add-backs = DOE',
            lambda c: f"{b.ref('ni', c)}+{b.sumrows(rl, c)}", lambda c: b.ref('doe', c),
            lambda p: (g(p, 'net_income') or 0) + sum(g(p, m) or 0 for m in RECON),
            lambda p: g(p, 'distributable_operating_earnings'))
    b.vals('  DOE — annuities', 'doe_annuities', indent=1)
    b.vals('  DOE — P&C', 'doe_pc', indent=1)
    b.vals('  DOE — life insurance', 'doe_life_insurance', indent=1)
    b.vals('  DOE — corporate & other', 'doe_corporate_and_other', indent=1)
    b.check('segment DOE foots',
            lambda c: b.sumrows(['  DOE — annuities', '  DOE — P&C', '  DOE — life insurance',
                                 '  DOE — corporate & other'], c),
            lambda c: b.ref('doe', c),
            lambda p: sum(g(p, m) or 0 for m in ('doe_annuities', 'doe_pc', 'doe_life_insurance',
                                                 'doe_corporate_and_other')),
            lambda p: g(p, 'distributable_operating_earnings'))

    b.section('⑥ STOCKS — balance sheet (FY2023 face balance sheet not printed in the 20-F: cells blank)')
    b.vals('Total assets', 'total_assets', mark='ta')
    b.vals('  total investments', 'total_investments', indent=1)
    b.vals("Policyholders' account balances", 'policyholders_account_balances', indent=1)
    b.vals('Future policy benefits', 'future_policy_benefits', indent=1)
    b.vals('Policy and contract claims', 'policy_and_contract_claims', indent=1)
    b.vals('Market risk benefits liability', 'market_risk_benefits_liability', indent=1)
    b.vals('Deposit liabilities', 'deposit_liabilities', indent=1)
    b.vals('Funds withheld for reinsurance liabilities', 'funds_withheld_for_reinsurance_liabilities', indent=1)
    b.vals('Separate account liabilities', 'separate_account_liabilities', indent=1)
    b.vals('Unearned premium reserve', 'unearned_premium_reserve', indent=1)
    b.vals('Total liabilities', 'total_liabilities', mark='tl')
    b.vals('Total equity', 'total_equity', mark='te')
    b.check('A = L + E',
            lambda c: f"{b.ref('tl', c)}+{b.ref('te', c)}", lambda c: b.ref('ta', c),
            lambda p: ((g(p, 'total_liabilities') + g(p, 'total_equity'))
                       if g(p, 'total_liabilities') is not None and g(p, 'total_equity') is not None else None),
            lambda p: g(p, 'total_assets'))
    b.vals('  of which non-controlling interests', 'non_controlling_interests_equity', indent=1)
    b.vals('Adjusted equity (non-GAAP, printed)', 'adjusted_equity')
    fy = b

    # ---- Engine (quarterly) ----
    q = Builder(wb, 'Engine', ['4Q24', '1Q25', '2Q25', '3Q25', '4Q25', '1Q26'], D)
    q.section('① THE SPREAD ENGINE — DOE basis (US$ millions)')
    q.vals('DOE net investment income', 'doe_net_investment_income', mark='qnii')
    q.vals('DOE cost of funds', 'doe_cost_of_funds', mark='qcof')
    q.vals('GROSS SPREAD (printed)', 'doe_gross_spread', mark='qgs')
    q.check('NII + cost of funds = gross spread',
            lambda c: f"{q.ref('qnii', c)}+{q.ref('qcof', c)}", lambda c: q.ref('qgs', c),
            lambda p: (g(p, 'doe_net_investment_income') or 0) + (g(p, 'doe_cost_of_funds') or 0),
            lambda p: g(p, 'doe_gross_spread'))
    q.vals('  interest expense', 'doe_interest_expense', indent=1)
    q.vals('  operating expenses and other', 'doe_operating_expenses_and_other', indent=1)
    q.vals('DISTRIBUTABLE OPERATING EARNINGS', 'distributable_operating_earnings', mark='qdoe')
    q.check('gross spread − interest − opex&other = DOE',
            lambda c: f"{q.ref('qgs', c)}+{q.sumrows(['  interest expense', '  operating expenses and other'], c)}",
            lambda c: q.ref('qdoe', c),
            lambda p: (g(p, 'doe_gross_spread') or 0) + sum(g(p, m) or 0 for m in
                       ('doe_interest_expense', 'doe_operating_expenses_and_other')),
            lambda p: g(p, 'distributable_operating_earnings'))
    q.vals('  income tax (segment-pretax view)', 'doe_income_tax_expense', indent=1, mark='qtax')
    q.vals('GAAP NET INCOME (LOSS)', 'gaap_net_income')
    q.note('GAAP vs DOE divergence: 1Q25 −282 vs +437 · 1Q26 −602 vs +438 — the sector pattern, quarterly.')
    q.vals('  pretax DOE — annuities', 'pretax_doe_annuities', indent=1)
    q.vals('  pretax DOE — P&C', 'pretax_doe_pc', indent=1)
    q.vals('  pretax DOE — life insurance', 'pretax_doe_life_insurance', indent=1)
    q.vals('  pretax DOE — corporate & other', 'pretax_doe_corporate_and_other', indent=1)
    q.vals('Total pretax DOE (printed)', 'total_pretax_doe', mark='qpre')
    q.check('pretax DOE + tax = DOE',
            lambda c: f"{q.ref('qpre', c)}+{q.ref('qtax', c)}", lambda c: q.ref('qdoe', c),
            lambda p: (g(p, 'total_pretax_doe') or 0) + (g(p, 'doe_income_tax_expense') or 0),
            lambda p: g(p, 'distributable_operating_earnings'))
    q.check('segment pretax DOE foots',
            lambda c: q.sumrows(['  pretax DOE — annuities', '  pretax DOE — P&C', '  pretax DOE — life insurance',
                                 '  pretax DOE — corporate & other'], c),
            lambda c: q.ref('qpre', c),
            lambda p: sum(g(p, m) or 0 for m in ('pretax_doe_annuities', 'pretax_doe_pc',
                                                 'pretax_doe_life_insurance', 'pretax_doe_corporate_and_other')),
            lambda p: g(p, 'total_pretax_doe'))
    q.section('② THE ANNUITY RULER — segment dollars (printed) and derived annualized rates')
    q.vals('Annuity net investment income', 'annuity_net_investment_income', mark='aNII')
    q.vals('Annuity cost of funds', 'annuity_cost_of_funds', mark='aCoF')
    q.vals('ANNUITY NET INVESTMENT SPREAD (printed)', 'annuity_net_investment_spread', mark='aNIS')
    q.check('NII + cost of funds = net spread',
            lambda c: f"{q.ref('aNII', c)}+{q.ref('aCoF', c)}", lambda c: q.ref('aNIS', c),
            lambda p: (g(p, 'annuity_net_investment_income') or 0) + (g(p, 'annuity_cost_of_funds') or 0),
            lambda p: g(p, 'annuity_net_investment_spread'))
    q.vals('Annuity average invested assets', 'annuity_average_invested_assets', mark='aAIA')
    q.formula('Earned rate (derived, ×4 / avg invested assets)',
              lambda c: f"=IF({q.ref('aAIA', c)}=0,\"\",{q.ref('aNII', c)}*4/{q.ref('aAIA', c)})", fmt=PCT)
    q.formula('Cost of funds rate (derived)',
              lambda c: f"=IF({q.ref('aAIA', c)}=0,\"\",{q.ref('aCoF', c)}*4/{q.ref('aAIA', c)})", fmt=PCT)
    q.formula('NET INVESTMENT SPREAD rate (derived)',
              lambda c: f"=IF({q.ref('aAIA', c)}=0,\"\",{q.ref('aNIS', c)}*4/{q.ref('aAIA', c)})", fmt=PCT)
    q.section('③ FLOWS — sales, outflows, net flows')
    q.vals('Retail sales — fixed index', 'retail_annuity_sales_fixed_index', indent=1)
    q.vals('Retail sales — fixed rate', 'retail_annuity_sales_fixed_rate', indent=1)
    q.vals('Retail sales — variable/other', 'retail_annuity_sales_variable_other', indent=1)
    q.vals('Total retail sales (printed)', 'total_retail_annuity_sales', mark='qretail')
    q.check('retail foots',
            lambda c: q.sumrows(['Retail sales — fixed index', 'Retail sales — fixed rate',
                                 'Retail sales — variable/other'], c),
            lambda c: q.ref('qretail', c),
            lambda p: sum(g(p, m) or 0 for m in ('retail_annuity_sales_fixed_index', 'retail_annuity_sales_fixed_rate',
                                                 'retail_annuity_sales_variable_other')),
            lambda p: g(p, 'total_retail_annuity_sales'))
    q.vals('Institutional — PRT', 'institutional_annuity_sales_prt', indent=1)
    q.vals('Institutional — funding agreements', 'institutional_annuity_sales_funding_agreements', indent=1)
    q.vals('Total institutional sales (printed)', 'total_institutional_annuity_sales', mark='qinst')
    q.check('institutional foots',
            lambda c: q.sumrows(['Institutional — PRT', 'Institutional — funding agreements'], c),
            lambda c: q.ref('qinst', c),
            lambda p: sum(g(p, m) or 0 for m in ('institutional_annuity_sales_prt',
                                                 'institutional_annuity_sales_funding_agreements')),
            lambda p: g(p, 'total_institutional_annuity_sales'))
    q.vals('TOTAL GROSS ANNUITY SALES (printed)', 'total_gross_annuity_sales', mark='qsales')
    q.check('retail + institutional = gross sales',
            lambda c: f"{q.ref('qretail', c)}+{q.ref('qinst', c)}", lambda c: q.ref('qsales', c),
            lambda p: (g(p, 'total_retail_annuity_sales') or 0) + (g(p, 'total_institutional_annuity_sales') or 0),
            lambda p: g(p, 'total_gross_annuity_sales'))
    q.vals('  retail outflows', 'retail_annuity_outflows', indent=1)
    q.vals('  institutional outflows', 'institutional_annuity_outflows', indent=1)
    q.vals('Total net annuity sales (printed)', 'total_net_annuity_sales')
    q.vals('TOTAL NET FLOWS (printed)', 'total_annuity_net_flows')
    q.section('④ INCOME ROWS (GAAP basis)')
    q.vals('Net premiums and other policy revenue', 'net_premiums_and_other_policy_revenue')
    q.vals('Net investment income incl funds withheld', 'net_investment_income_incl_funds_withheld')
    q.vals('Net investment gains (losses) incl funds withheld', 'net_investment_gains_losses_incl_funds_withheld')
    q.section('⑤ STOCKS')
    q.vals('Total assets', 'total_assets', mark='qta')
    q.vals('  total insurance assets', 'total_insurance_assets', indent=1)
    q.vals('Total liabilities', 'total_liabilities', mark='qtl')
    q.vals('  total insurance liabilities', 'total_insurance_liabilities', indent=1)
    q.vals('  total gross reserves', 'total_gross_reserves', indent=1)
    q.vals('  total net reserves', 'total_net_reserves', indent=1)
    q.vals('Total equity', 'total_equity', mark='qte')
    q.check('A = L + E',
            lambda c: f"{q.ref('qtl', c)}+{q.ref('qte', c)}", lambda c: q.ref('qta', c),
            lambda p: ((g(p, 'total_liabilities') + g(p, 'total_equity'))
                       if g(p, 'total_liabilities') is not None and g(p, 'total_equity') is not None else None),
            lambda p: g(p, 'total_assets'))
    q.vals('Adjusted equity (non-GAAP)', 'adjusted_equity')
    q.section('⑥ CROSS-GATE — 2025 quarters vs FY2025 (Engine-FY column D)')
    for lab, met, fymark in (('net income', 'gaap_net_income', 'ni'),
                             ('DOE', 'distributable_operating_earnings', 'doe'),
                             ('gross annuity sales', 'total_gross_annuity_sales', 'sales_tot')):
        fyrow = fy.marks[fymark]
        q.row += 1
        q.ws.cell(row=q.row, column=1, value=f'CHECK: Σ(1Q25..4Q25) {lab} = FY2025').font = GREY
        # quarter columns C..F are 1Q25..4Q25
        src_row = {'gaap_net_income': q.marks['GAAP NET INCOME (LOSS)'],
                   'distributable_operating_earnings': q.marks['qdoe'],
                   'total_gross_annuity_sales': q.marks['qsales']}[met]
        f = (f"=IF(ABS(SUM(C{src_row}:F{src_row})-'Engine-FY'!D{fyrow})<=2,\"OK\",\"FAIL\")")
        c = q.ws.cell(row=q.row, column=2, value=f)
        c.font = GREY
        q.checks.append((q.ws.title, f'B{q.row}'))
        qs = sum(g(p, met) or 0 for p in ('1Q25', '2Q25', '3Q25', '4Q25'))
        fyv = g('FY2025', {'gaap_net_income': 'net_income',
                           'distributable_operating_earnings': 'distributable_operating_earnings',
                           'total_gross_annuity_sales': 'gross_annuity_sales_total'}[met])
        if fyv is not None and abs(qs - fyv) > 2:
            q.py_fail.append(f'cross-gate {lab}: quarters {qs:,.0f} vs FY {fyv:,.0f}')

    readme(wb, [
        'Brookfield Wealth Solutions (BNT) — quarterly + annual engine',
        '',
        'Units: US$ millions as printed. Blue = value as printed in the source; black bold = live formula;',
        'CHECK rows verify identities cell-by-cell (OK / FAIL / n/a where a period is not published).',
        'Checks tab = master count of failing checks (must read 0).',
        '',
        'Sources: BNT 20-F FY2025 (three-year statements; FY2023 face balance sheet not printed — cells blank,',
        'boundary logged) · BWS supplementals Q4-2025 + Q1-2026 (quarterly series began 1Q25; 4Q24 comparative).',
        'All documents sha-tracked in acquisition/manifest.csv; extraction in extract/cross-section/bnt_engine_raw.csv',
        '(481 rows, anchor-quoted). Regenerate: python3 tools/build_platform_engines.py bnt',
        '',
        'Boundary: no quarterly series exists before 1Q25 (BNT is a foreign private issuer, 20-F/6-K cadence).',
        'The BN/BNT merger (approved 2026-07-16) is expected to end standalone reporting.',
    ])
    all_checks = fy.checks + q.checks
    checks_sheet(wb, all_checks, 'BNT engine.')
    dest = ROOT / 'dossiers/brookfield/bnt-quarterly-engine.xlsx'
    wb.save(dest)
    fails = fy.py_fail + q.py_fail
    print(f'wrote {dest} ({len(all_checks)} checks; python-verified, {len(fails)} FAIL)')
    for f_ in fails:
        print('  PY-FAIL:', f_)
    return not fails


# ---------------------------------------------------------------- Ares/Aspida
def build_aspida():
    D = load('aspida_engine_raw', ['485bpos', 'n4a'])
    wb = Workbook()
    wb.remove(wb.active)
    g = lambda p, m: D.get((p, m))
    b = Builder(wb, 'Engine-FY', ['FY2023', 'FY2024', 'FY2025'], D)

    b.section('① MONEY IN — premiums and deposits (US$ thousands, statutory basis)')
    b.vals('Premiums — direct (note 5)', 'note5_premiums_earned_direct', indent=1)
    b.vals('Premiums — reinsurance assumed (note 5)', 'note5_premiums_earned_reinsurance_assumed', indent=1)
    b.vals('Premiums — reinsurance ceded (note 5)', 'note5_premiums_earned_reinsurance_ceded', indent=1)
    b.vals('Net premiums earned (note 5)', 'note5_net_premiums_earned', mark='n5net')
    b.check('direct + assumed + ceded(signed) = net (note 5)',
            lambda c: b.sumrows(['Premiums — direct (note 5)', 'Premiums — reinsurance assumed (note 5)',
                                 'Premiums — reinsurance ceded (note 5)'], c),
            lambda c: b.ref('n5net', c),
            lambda p: (g(p, 'note5_premiums_earned_direct') or 0) + (g(p, 'note5_premiums_earned_reinsurance_assumed') or 0)
            + (g(p, 'note5_premiums_earned_reinsurance_ceded') or 0),
            lambda p: g(p, 'note5_net_premiums_earned'))
    b.vals('Premium & annuity considerations, net (face)', 'ops_premium_and_annuity_considerations_life_net', mark='faceprem')
    b.check('note 5 net = face premium line',
            lambda c: b.ref('n5net', c), lambda c: b.ref('faceprem', c),
            lambda p: g(p, 'note5_net_premiums_earned'),
            lambda p: g(p, 'ops_premium_and_annuity_considerations_life_net'))
    b.vals('Net deposits on deposit-type contracts (cash flow)', 'cf_net_deposits_on_deposit_type_contracts', mark='deps')
    b.formula('TOTAL INFLOWS (net premiums + deposits)',
              lambda c: f"={b.ref('faceprem', c)}+{b.ref('deps', c)}")
    b.note('Funding agreements launched 2025: deposit-type liability 22 → 8,041 → 819,312.')

    b.section('② REVENUES (statutory summary of operations)')
    REV = [('Premium & annuity considerations, net', 'ops_premium_and_annuity_considerations_life_net'),
           ('Considerations — supplementary contracts', 'ops_considerations_supplementary_contracts_life_contingencies'),
           ('Net investment income', 'ops_net_investment_income'),
           ('Amortization of IMR', 'ops_amortization_of_imr'),
           ('Commissions & expense allowances on reinsurance ceded', 'ops_commissions_expense_allowances_reinsurance_ceded'),
           ('Modco assumed adjustment', 'ops_modco_assumed_adjustment'),
           ('Amortization of deferred reinsurance gain', 'ops_amortization_deferred_reinsurance_gain'),
           ('Separate accounts net gain', 'ops_separate_accounts_net_gain_from_operations'),
           ('Miscellaneous income', 'ops_miscellaneous_income')]
    for lab, m in REV:
        b.vals(lab, m, indent=1)
    b.vals('TOTAL PREMIUMS AND OTHER REVENUES (printed)', 'ops_total_premiums_and_other_revenues', mark='rev_tot')
    b.check('revenue lines foot',
            lambda c: b.sumrows([lab for lab, _ in REV], c), lambda c: b.ref('rev_tot', c),
            lambda p: sum(g(p, m) or 0 for _, m in REV),
            lambda p: g(p, 'ops_total_premiums_and_other_revenues'))

    b.section('③ BENEFITS AND OTHER DEDUCTIONS')
    DED = [('Annuity benefits', 'ops_annuity_benefits'),
           ('Surrender benefits and withdrawals', 'ops_surrender_benefits_and_withdrawals'),
           ('Interest on deposit-type contract funds', 'ops_interest_adjustments_contract_deposit_type_funds'),
           ('Increase in aggregate reserves', 'ops_increase_in_aggregate_reserves_life'),
           ('Net transfers to/(from) separate accounts', 'ops_net_transfers_to_from_separate_accounts'),
           ('Commissions and brokerage expense', 'ops_commissions_and_brokerage_expense'),
           ('Commission expense allowance on reinsurance assumed', 'ops_commission_expense_allowance_reinsurance_assumed'),
           ('Reinsurance investment credit', 'ops_reinsurance_investment_credit'),
           ('General insurance expenses', 'ops_general_insurance_expenses'),
           ('Insurance taxes, licenses and fees', 'ops_insurance_taxes_licenses_fees')]
    for lab, m in DED:
        b.vals(lab, m, indent=1)
    b.vals('TOTAL BENEFITS AND OTHER DEDUCTIONS (printed)', 'ops_total_benefits_and_other_deductions', mark='ded_tot')
    b.check('deduction lines foot',
            lambda c: b.sumrows([lab for lab, _ in DED], c), lambda c: b.ref('ded_tot', c),
            lambda p: sum(g(p, m) or 0 for _, m in DED),
            lambda p: g(p, 'ops_total_benefits_and_other_deductions'))

    b.section('④ NET INCOME CHAIN')
    b.vals('Loss from operations before FIT (printed)', 'ops_loss_from_operations_before_fit', mark='prefit')
    b.check('revenues − deductions = pre-FIT',
            lambda c: f"{b.ref('rev_tot', c)}-{b.ref('ded_tot', c)}", lambda c: b.ref('prefit', c),
            lambda p: (g(p, 'ops_total_premiums_and_other_revenues') or 0) - (g(p, 'ops_total_benefits_and_other_deductions') or 0),
            lambda p: g(p, 'ops_loss_from_operations_before_fit'))
    b.vals('Federal income tax (expense) benefit', 'ops_fit_expense_benefit', mark='fit')
    b.vals('Net loss from operations before realized gains (printed)', 'ops_net_loss_from_operations_before_realized_gains', mark='prereal')
    b.check('pre-FIT − FIT = pre-realized',
            lambda c: f"{b.ref('prefit', c)}-{b.ref('fit', c)}", lambda c: b.ref('prereal', c),
            lambda p: (g(p, 'ops_loss_from_operations_before_fit') or 0) - (g(p, 'ops_fit_expense_benefit') or 0),
            lambda p: g(p, 'ops_net_loss_from_operations_before_realized_gains'))
    b.vals('Net realized investment gains, net of tax', 'ops_net_realized_investment_gains_net_of_tax', mark='realized')
    b.vals('NET INCOME (LOSS) (printed)', 'ops_net_income_loss', mark='ni')
    b.check('pre-realized + realized = net income',
            lambda c: f"{b.ref('prereal', c)}+{b.ref('realized', c)}", lambda c: b.ref('ni', c),
            lambda p: (g(p, 'ops_net_loss_from_operations_before_realized_gains') or 0) + (g(p, 'ops_net_realized_investment_gains_net_of_tax') or 0),
            lambda p: g(p, 'ops_net_income_loss'))
    b.note('Source discrepancy, logged: FY2023 net loss prints (48,758) in operations vs (48,760) in the C&S'
           ' rollforward — $2k as printed in both filings; the rollforward foots with (48,760).')

    b.section('⑤ CAPITAL & SURPLUS ROLLFORWARD')
    b.vals('C&S — beginning of year', 'cs_beginning_balance', mark='cs0')
    ROLL = [('Net loss (per C&S statement)', 'cs_net_loss'),
            ('Change in net unrealized capital gains', 'cs_change_net_unrealized_capital_gains'),
            ('Change in nonadmitted assets', 'cs_change_nonadmitted_assets'),
            ('Change in asset valuation reserve', 'cs_change_asset_valuation_reserve'),
            ('Change in net deferred income tax', 'cs_change_net_deferred_income_tax'),
            ('Capital contributions', 'cs_capital_contributions'),
            ('Ceding unrealized gains — reinsurance', 'cs_ceding_unrealized_gains_losses_reinsurance'),
            ('Deferred reinsurance gain, net of amortization', 'cs_deferred_reinsurance_gain_net_of_amortization')]
    for lab, m in ROLL:
        b.vals(lab, m, indent=1)
    b.vals('C&S — END OF YEAR (printed)', 'cs_ending_balance', mark='cs1')
    b.check('rollforward foots',
            lambda c: f"{b.ref('cs0', c)}+{b.sumrows([lab for lab, _ in ROLL], c)}",
            lambda c: b.ref('cs1', c),
            lambda p: (g(p, 'cs_beginning_balance') or 0) + sum(g(p, m) or 0 for _, m in ROLL),
            lambda p: g(p, 'cs_ending_balance'))

    b.section('⑥ STOCKS — balance sheet')
    b.vals('Total admitted assets', 'bs_total_admitted_assets', mark='ta')
    b.vals('  total cash and invested assets', 'bs_total_cash_and_invested_assets', indent=1)
    b.vals('Total liabilities', 'bs_total_liabilities', mark='tl')
    b.vals('  aggregate reserve for life contracts (net)', 'bs_aggregate_reserve_life_contracts', indent=1)
    b.vals('  liability for deposit-type contracts', 'bs_liability_deposit_type_contracts', indent=1)
    b.vals('  funds held under reinsurance treaties (unauthorized)', 'bs_funds_held_under_reinsurance_treaties_unauthorized', indent=1)
    b.vals('  interest maintenance reserve', 'bs_interest_maintenance_reserve_liability', indent=1)
    b.vals('  asset valuation reserve', 'bs_asset_valuation_reserve', indent=1)
    b.vals('TOTAL CAPITAL AND SURPLUS (printed)', 'bs_total_capital_and_surplus', mark='cs_bs')
    b.check('A = L + C&S',
            lambda c: f"{b.ref('tl', c)}+{b.ref('cs_bs', c)}", lambda c: b.ref('ta', c),
            lambda p: (g(p, 'bs_total_liabilities') or 0) + (g(p, 'bs_total_capital_and_surplus') or 0),
            lambda p: g(p, 'bs_total_admitted_assets'))
    b.check('C&S rollforward end = balance sheet C&S',
            lambda c: b.ref('cs1', c), lambda c: b.ref('cs_bs', c),
            lambda p: g(p, 'cs_ending_balance'), lambda p: g(p, 'bs_total_capital_and_surplus'))
    b.section('⑦ THE REINSURANCE MACHINE — gross vs ceded reserves (note 6)')
    b.vals('Gross annuity reserves', 'note6_gross_annuity_reserves_total', mark='rgross')
    b.vals('Ceded annuity reserves', 'note6_ceded_annuity_reserves', mark='rceded')
    b.vals('Net annuity reserves (printed)', 'note6_net_annuity_reserves_total', mark='rnet')
    b.check('gross − ceded = net reserves',
            lambda c: f"{b.ref('rgross', c)}-{b.ref('rceded', c)}", lambda c: b.ref('rnet', c),
            lambda p: (g(p, 'note6_gross_annuity_reserves_total') or 0) - (g(p, 'note6_ceded_annuity_reserves') or 0),
            lambda p: g(p, 'note6_net_annuity_reserves_total'))
    b.formula('CEDED SHARE OF GROSS RESERVES', lambda c: f"=IF({b.ref('rgross', c)}=0,\"\",{b.ref('rceded', c)}/{b.ref('rgross', c)})", fmt=PCT)
    b.note('21.1% → 57.9% → 60.6% — more than half the machine now runs through reinsurance counterparties.')

    readme(wb, [
        'Aspida Life Insurance Company — annual engine (statutory basis)',
        '',
        'Units: US$ thousands as printed. Blue = printed value; black bold = live formula; CHECK rows',
        'verify identities (OK/FAIL/n/a). Checks tab = master count of failing checks (must read 0).',
        '',
        'Sources: EY-audited statutory financial statements embedded in EDGAR RILA filings (CIK 1934234):',
        '485BPOS filed 2026 (FY2025+FY2024) and N-4 filed 2025 (FY2023). Extraction:',
        'extract/cross-section/aspida_engine_raw.csv (257 rows, anchor-quoted).',
        'Regenerate: python3 tools/build_platform_engines.py aspida',
        '',
        'Boundary: NO quarterly cadence exists anywhere in public documents (Ares 10-Qs carry zero Aspida',
        'data; no NAIC blanks are published). This annual engine is the maximum public resolution.',
    ])
    checks_sheet(wb, b.checks, 'Aspida engine.')
    dest = ROOT / 'dossiers/ares/aspida-annual-engine.xlsx'
    wb.save(dest)
    print(f'wrote {dest} ({len(b.checks)} checks; python-verified, {len(b.py_fail)} FAIL)')
    for f_ in b.py_fail:
        print('  PY-FAIL:', f_)
    return not b.py_fail


# ---------------------------------------------------------------- Blue Owl/KLR
def load_klr():
    """Strict single-source per column (extraction-agent confirmed compositions):
    FY2022/FY2023 <- klr-fs-2023 (pre-LDTI); FY2024/FY2025 <- klr-fs-2025 (LDTI restated).
    No cross-source fallback — a metric absent from the designated source stays blank."""
    src_for = {'FY2022': 'klr-fs-2023', 'FY2023': 'klr-fs-2023',
               'FY2024': 'klr-fs-2025', 'FY2025': 'klr-fs-2025'}
    out = {}
    for r in csv.DictReader(open(ROOT / 'extract/cross-section/klr_engine_raw.csv')):
        p = r['period']
        if p in src_for and src_for[p] in r['source_file']:
            out[(p, r['metric'])] = float(r['value'])
    return out


def build_klr():
    D = load_klr()
    wb = Workbook()
    wb.remove(wb.active)
    g = lambda p, m: D.get((p, m))
    b = Builder(wb, 'Engine-FY', ['FY2022', 'FY2023', 'FY2024', 'FY2025'], D)

    b.section('① REVENUES (US dollars as printed; FY2022/23 pre-LDTI, FY2024/25 LDTI-restated basis)')
    REV = [('Premium income', 'premium_income'),
           ('Investment income — funds withheld', 'investment_income_funds_withheld'),
           ('Realized losses — funds withheld', 'realized_losses_funds_withheld'),
           ('Derivative gains (losses) — funds withheld', 'derivative_gains_losses_funds_withheld'),
           ('Unrealized gains (losses) — embedded derivative', 'unrealized_gains_losses_embedded_derivative'),
           ('Amortization of deferred profit liability (pre-LDTI)', 'amortization_deferred_profit_liability'),
           ('Amortization of unearned revenue reserve (LDTI)', 'amortization_unearned_revenue_reserve'),
           ('Investment management expenses (contra)', 'investment_management_expenses'),
           ('Net investment income', 'net_investment_income'),
           ('Realized gains (losses) — investments & derivatives', 'realized_gains_losses_investments_derivatives'),
           ('Unrealized gains (losses) — investments', 'unrealized_gains_losses_investments'),
           ('Unrealized losses — derivatives', 'unrealized_losses_derivatives')]
    for lab, m in REV:
        b.vals(lab, m, indent=1)
    b.vals('TOTAL REVENUES (printed)', 'total_revenues', mark='rev_tot')
    b.check('revenue lines foot',
            lambda c: b.sumrows([lab for lab, _ in REV], c), lambda c: b.ref('rev_tot', c),
            lambda p: sum(g(p, m) or 0 for _, m in REV), lambda p: g(p, 'total_revenues'))

    b.section('② BENEFITS AND EXPENSES (no printed total in any year — computed; signed as stored)')
    EXP = [('Claims and other insurance expenses', 'claims_and_other_insurance_expenses'),
           ('Interest credited', 'interest_credited'),
           ('Change in fair value of market risk benefits (LDTI)', 'change_in_fair_value_of_mrb_net'),
           ('Amortization of deferred acquisition costs', 'amortization_deferred_acquisition_costs'),
           ('Amortization of deferred commission (pre-LDTI)', 'amortization_deferred_commission'),
           ('Net foreign exchange (gains) losses', 'net_foreign_exchange_gains_losses'),
           ('Operating expenses', 'operating_expenses')]
    for lab, m in EXP:
        b.vals(lab, m, indent=1)
    b.formula('TOTAL BENEFITS AND EXPENSES (computed)',
              lambda c: '=' + b.sumrows([lab for lab, _ in EXP], c), mark='exp_tot')

    b.section('③ NET INCOME AND COMPREHENSIVE INCOME')
    b.vals('NET INCOME (LOSS) (printed)', 'net_income_loss', mark='ni')
    b.check('revenues − expenses = net income',
            lambda c: f"{b.ref('rev_tot', c)}-{b.ref('exp_tot', c)}", lambda c: b.ref('ni', c),
            lambda p: (g(p, 'total_revenues') or 0) - sum(g(p, m) or 0 for _, m in EXP),
            lambda p: g(p, 'net_income_loss'))
    b.vals('Total other comprehensive income (loss)', 'total_other_comprehensive_income_loss', mark='oci')
    b.vals('TOTAL COMPREHENSIVE INCOME (LOSS) (printed)', 'total_comprehensive_income_loss', mark='ci')
    b.check('NI + OCI = comprehensive income',
            lambda c: f"{b.ref('ni', c)}+{b.ref('oci', c)}", lambda c: b.ref('ci', c),
            lambda p: (g(p, 'net_income_loss') or 0) + (g(p, 'total_other_comprehensive_income_loss') or 0),
            lambda p: g(p, 'total_comprehensive_income_loss'))

    b.section('④ STOCKS — balance sheet')
    b.vals('Total cash and invested assets', 'total_cash_and_invested_assets', indent=1)
    b.vals('  fixed maturities AFS', 'fixed_maturity_afs', indent=2)
    b.vals('  fixed maturities trading', 'fixed_maturity_trading', indent=2)
    b.vals('Funds withheld assets', 'funds_withheld_assets', indent=1)
    b.vals('Reinsurance recoverable', 'reinsurance_recoverable', indent=1)
    b.vals('Deferred acquisition costs', 'deferred_acquisition_costs', indent=1)
    b.vals('TOTAL ASSETS (printed)', 'total_assets', mark='ta')
    b.vals('Future policy benefit reserves (pre-LDTI)', 'future_policy_benefit_reserves', indent=1)
    b.vals('Policyholder account balance (LDTI)', 'policyholder_account_balance', indent=1)
    b.vals('Liability for future policy benefits (LDTI)', 'liability_for_future_policy_benefits', indent=1)
    b.vals('Market risk benefit liability (LDTI)', 'market_risk_benefit_liability', indent=1)
    b.vals('Unearned revenue reserve (LDTI)', 'unearned_revenue_reserve', indent=1)
    b.vals('Deposit liability', 'deposit_liability', indent=1)
    b.vals('Funds withheld liabilities', 'funds_withheld_liabilities', indent=1)
    b.vals('Repurchase agreement liabilities', 'repurchase_agreement_liabilities', indent=1)
    b.vals('Embedded derivative liabilities', 'embedded_derivative_liabilities', indent=1)
    b.vals('Due to affiliates', 'due_to_affiliates', indent=1)
    b.vals('TOTAL LIABILITIES (printed)', 'total_liabilities', mark='tl')
    b.vals('  share capital', 'share_capital', indent=1)
    b.vals('  additional paid-in capital', 'additional_paid_in_capital', indent=1)
    b.vals('  retained earnings (deficit)', 'retained_earnings_deficit', indent=1)
    b.vals('  accumulated other comprehensive income', 'accumulated_other_comprehensive_income', indent=1)
    b.vals("TOTAL SHAREHOLDER'S EQUITY (printed)", 'total_shareholders_equity', mark='te')
    b.check('equity components foot',
            lambda c: b.sumrows(['  share capital', '  additional paid-in capital',
                                 '  retained earnings (deficit)', '  accumulated other comprehensive income'], c),
            lambda c: b.ref('te', c),
            lambda p: sum(g(p, m) or 0 for m in ('share_capital', 'additional_paid_in_capital',
                                                 'retained_earnings_deficit', 'accumulated_other_comprehensive_income')),
            lambda p: g(p, 'total_shareholders_equity'))
    b.check('A = L + E',
            lambda c: f"{b.ref('tl', c)}+{b.ref('te', c)}", lambda c: b.ref('ta', c),
            lambda p: (g(p, 'total_liabilities') or 0) + (g(p, 'total_shareholders_equity') or 0),
            lambda p: g(p, 'total_assets'))
    b.formula('EQUITY / ASSETS', lambda c: f"=IF({b.ref('ta', c)}=0,\"\",{b.ref('te', c)}/{b.ref('ta', c)})", fmt=PCT)
    b.note('Equity was NEGATIVE at YE2022 (−$174.6M) and YE2023 (−$64.1M); rebuilt to $171.9M by YE2025'
           ' against $9.8B of assets. FY2024 column is the ASU 2018-12 (LDTI) restated basis per klr-fs-2025;'
           ' as-reported FY2024 (net loss −$37.2M, equity +$7.7M) is preserved in the extraction CSV.')

    readme(wb, [
        'Kuvare Life Re Ltd. (Bermuda) — annual engine (audited GAAP financial statements)',
        '',
        'Units: US DOLLARS as printed (not thousands). Blue = printed value; black bold = live formula;',
        'CHECK rows verify identities (OK/FAIL/n/a). Checks tab = master count (must read 0).',
        '',
        'Column basis is single-source, never mixed: FY2022/FY2023 from klr-fs-2023.pdf (pre-LDTI);',
        'FY2024/FY2025 from klr-fs-2025.pdf (ASU 2018-12 restated). Pre-/post-LDTI line differences render',
        'as blank cells on the inapplicable basis (boundary, not omission).',
        'Extraction: extract/cross-section/klr_engine_raw.csv (374 rows, anchor-quoted).',
        'Regenerate: python3 tools/build_platform_engines.py klr',
        '',
        'Source print error, logged: klr-fs-2025 FY2024 cash-flow comparative prints operating cash 1,601,181,219',
        'where its own components and restatement note prove 1,681,181,219 (an $80M misprint).',
        'Boundary: KLR is the ONLY public financial-statement window into the Kuvare group; the three US',
        'carriers publish no statutory statements and no Kuvare GAAP exists publicly.',
    ])
    checks_sheet(wb, b.checks, 'KLR engine.')
    dest = ROOT / 'dossiers/blueowl/klr-annual-engine.xlsx'
    wb.save(dest)
    print(f'wrote {dest} ({len(b.checks)} checks; python-verified, {len(b.py_fail)} FAIL)')
    for f_ in b.py_fail:
        print('  PY-FAIL:', f_)
    return not b.py_fail


if __name__ == '__main__':
    want = sys.argv[1:] or ['bnt', 'aspida', 'klr']
    ok = True
    for w in want:
        r = {'bnt': build_bnt, 'aspida': build_aspida, 'klr': build_klr}[w]()
        ok = ok and (r is not False)
    sys.exit(0 if ok else 1)
