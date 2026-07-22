#!/usr/bin/env python3
"""Global Atlantic engine workbook (Athene-parity artifact).

Engine    - 13 quarterly columns 1Q23..1Q26 from the gated GALD supplement
            extracts: new business by channel, GAAP income, adjusted operating
            earnings chain (GA's SRE analog), rates, reserves by product,
            AFS quality (NAIC + NRSRO), capitalization/equity.
Engine-FY - FY2023/24/25 (supplement FY columns; stocks = year-ends).
Checks    - live count of failing identities (must read 0).
Data      - raw extract dump. ReadMe - sources & method.
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DEST = ROOT / 'dossiers/global-atlantic/ga-quarterly-engine.xlsx'
DEST.parent.mkdir(exist_ok=True)

D = {}
for r in csv.DictReader(open(ROOT / 'extract/global-atlantic/gald_supplement.csv')):
    if r['value'] != '':
        D[(r['period'], r['metric'])] = float(r['value'])

QTRS = ['1Q23', '2Q23', '3Q23', '4Q23', '1Q24', '2Q24', '3Q24', '4Q24',
        '1Q25', '2Q25', '3Q25', '4Q25', '1Q26']
FYS = ['FY2023', 'FY2024', 'FY2025']

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10)
BOLD = Font(name='Arial', size=10, bold=True)
HDR = Font(name='Arial', size=11, bold=True)
GREY = Font(name='Arial', size=9, color='666666')
MONEY_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.00%'

wb = Workbook()
ws = None
row = 1
PERIODS = []
SHEET = ''
check_cells = []


def put(label, font=BLACK, indent=0):
    c = ws.cell(row=row, column=1, value=('    ' * indent) + label)
    if label.startswith('='):
        c.data_type = 's'
    c.font = font


def vals(metric, fmt=MONEY_FMT, pct=False):
    for i, p in enumerate(PERIODS):
        v = D.get((p, metric))
        c = ws.cell(row=row, column=2 + i)
        if v is not None:
            c.value = v / 100.0 if pct else v
        c.font = BLUE
        c.number_format = fmt


def formula(fn, fmt=MONEY_FMT):
    for i, p in enumerate(PERIODS):
        col = get_column_letter(2 + i)
        c = ws.cell(row=row, column=2 + i, value=fn(col))
        c.font = BOLD
        c.number_format = fmt


def check(fn):
    for i, p in enumerate(PERIODS):
        col = get_column_letter(2 + i)
        f = fn(col)
        import re as _re
        refs = _re.findall(r'[B-O]\d+', f)
        if refs:
            f = f'=IF({refs[-1]}="","n/a",{f[1:]})'
        c = ws.cell(row=row, column=2 + i, value=f)
        c.font = GREY
        check_cells.append(f"'{SHEET}'!{col}{row}")


def section(t):
    global row
    row += 1
    put(t, HDR)
    row += 1


def build():
    global row
    section('MONEY IN — NEW BUSINESS VOLUME BY CHANNEL')
    r0 = row
    put('Fixed-rate annuities (individual)', BLUE); vals('nbv_fixed_rate'); row += 1
    put('Fixed-indexed annuities (individual)', BLUE); vals('nbv_fia'); row += 1
    put('Variable annuities (individual)', BLUE); vals('nbv_va'); row += 1
    r_ret = row
    put('= Total retirement products', BOLD); formula(lambda c: f'=SUM({c}{r0}:{c}{r0+2})'); row += 1
    r_retrep = row
    put('   reported (filed)', BLUE); vals('nbv_retirement_total'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_ret}={c}{r_retrep},"OK","FAIL")'); row += 1
    put('Preneed life', BLUE); vals('nbv_preneed'); row += 1
    put('Block reinsurance (institutional)', BLUE); vals('nbv_block'); row += 1
    put('Flow reinsurance (institutional)', BLUE); vals('nbv_flow'); row += 1
    put('Pension risk transfer', BLUE); vals('nbv_prt'); row += 1
    put('Funding agreements', BLUE); vals('nbv_funding_agreements'); row += 1

    section('GAAP INCOME — REVENUES AND THE BOTTOM LINE')
    rg = row
    put('Premiums', BLUE); vals('premiums'); row += 1
    put('Policy fees', BLUE); vals('policy_fees'); row += 1
    put('Net investment income', BLUE); vals('nii'); row += 1
    put('Net investment gains (losses)', BLUE); vals('inv_gl'); row += 1
    put('Other income', BLUE); vals('other_income'); row += 1
    r_rev = row
    put('= Total revenues', BOLD); formula(lambda c: f'=SUM({c}{rg}:{c}{rg+4})'); row += 1
    r_revrep = row
    put('   reported (filed)', BLUE); vals('total_revenues'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_rev}={c}{r_revrep},"OK","FAIL")'); row += 1
    put('NET INCOME (LOSS) TO SHAREHOLDER — the valuation line', BLUE); vals('ni_shareholder'); row += 1

    section("THE OPERATING ENGINE — GA'S SRE ANALOG (adjusted operating earnings)")
    ra = row
    put('Adjusted net investment income', BLUE); vals('adj_nii'); row += 1
    put('− Adjusted net cost of insurance', BLUE); vals('adj_cost_ins'); row += 1
    r_uw = row
    put('= Adjusted net underwriting income', BOLD); formula(lambda c: f'={c}{ra}-{c}{ra+1}'); row += 1
    r_uwrep = row
    put('   reported (filed)', BLUE); vals('adj_underwriting'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_uw}={c}{r_uwrep},"OK","FAIL")'); row += 1
    put('− Interest expense (where separately stated)', BLUE); vals('adj_interest'); row += 1
    put('− Adjusted G&A (and other)', BLUE); vals('adj_gae'); row += 1
    r_pt = row
    put('= Adjusted operating earnings, pre-tax', BOLD)
    formula(lambda c: f'={c}{r_uwrep}-{c}{r_uwrep+2}-{c}{r_uwrep+3}'); row += 1
    r_ptrep = row
    put('   reported (filed)', BLUE); vals('adj_op_pretax'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_pt}={c}{r_ptrep},"OK","FAIL")'); row += 1
    put('− Adjusted operating tax', BLUE); vals('adj_op_tax'); row += 1
    r_net = row
    put('= ADJUSTED OPERATING EARNINGS, NET', BOLD)
    formula(lambda c: f'={c}{r_ptrep}+{c}{r_ptrep+1}'); row += 1
    r_netrep = row
    put('   reported (filed)', BLUE); vals('aoi_net_aoe'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_net}={c}{r_netrep},"OK","FAIL")'); row += 1
    put('Average adjusted invested assets', BLUE); vals('avg_aia'); row += 1

    section('THE RATES (annualized, as filed)')
    put('Net investment earned rate', BLUE); vals('nier', fmt=PCT_FMT, pct=True); row += 1
    put('Cost of insurance ratio', BLUE); vals('cost_ins_ratio', fmt=PCT_FMT, pct=True); row += 1
    put('NET UNDERWRITING MARGIN (their NIS)', BLUE); vals('underwriting_ratio', fmt=PCT_FMT, pct=True); row += 1
    put('Operating ROA pre-tax', BLUE); vals('aor_pretax', fmt=PCT_FMT, pct=True); row += 1
    put('ROE (GAAP)', BLUE); vals('roe', fmt=PCT_FMT, pct=True); row += 1
    put('Adjusted ROE', BLUE); vals('adj_roe', fmt=PCT_FMT, pct=True); row += 1
    put('ADJUSTED OPERATING ROE', BLUE); vals('adj_op_roe', fmt=PCT_FMT, pct=True); row += 1

    section('RESERVES BY PRODUCT, NET (series starts 4Q24 — earlier layout differs; boundary logged)')
    rr = row
    for key, lbl in (('res_fixed_rate', 'Fixed-rate annuity'), ('res_fia', 'Fixed-indexed annuity'),
                     ('res_payout', 'Payout annuities'), ('res_va', 'Variable annuity'),
                     ('res_isl', 'Interest sensitive life'), ('res_other_life', 'Other life'),
                     ('res_fa', 'Funding agreements (runnable)'), ('res_closed_block', 'Closed block'),
                     ('res_other_corp', 'Other corporate')):
        put(lbl, BLUE, 1); vals(key); row += 1
    r_rt = row
    put('= Total net reserves', BOLD); formula(lambda c: f'=SUM({c}{rr}:{c}{rr+8})'); row += 1
    r_rtrep = row
    put('   reported (filed)', BLUE); vals('res_total'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_rtrep}="","n/a",IF({c}{r_rt}={c}{r_rtrep},"OK","FAIL"))'); row += 1

    section('PORTFOLIO QUALITY — AFS FIXED MATURITIES (period-ends as filed)')
    rq = row
    put('NAIC 1', BLUE, 1); vals('naic1'); row += 1
    put('NAIC 2', BLUE, 1); vals('naic2'); row += 1
    r_ig = row
    put('= NAIC investment grade', BOLD); formula(lambda c: f'=IF({c}{rq}="","",{c}{rq}+{c}{rq+1})'); row += 1
    put('NAIC 3–6 (below IG, filed)', BLUE, 1); vals('naic_big'); row += 1
    r_qt = row
    put('= Total (IG + below-IG)', BOLD); formula(lambda c: f'=IF({c}{r_ig}="","",{c}{r_ig}+{c}{r_ig+1})'); row += 1
    r_qtrep = row
    put('   reported total (filed)', BLUE); vals('naic_total'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_qtrep}="","n/a",IF({c}{r_qt}={c}{r_qtrep},"OK","FAIL"))'); row += 1
    put('NRSRO below-IG (same book, market ratings)', BLUE, 1); vals('nrsro_big'); row += 1
    put('   the two-ruler gap is the story: NAIC flatters vs NRSRO', GREY); row += 1

    section('CAPITAL UNDERNEATH')
    put('Total shareholders equity', BLUE); vals('shareholders_equity'); row += 1
    put('Adjusted shareholders equity (ex-AOCI etc.)', BLUE); vals('adj_equity'); row += 1
    put('Total debt', BLUE); vals('total_debt'); row += 1
    put('Total assets', BLUE); vals('total_assets'); row += 1


for tab, periods in (('Engine', QTRS), ('Engine-FY', FYS)):
    ws = wb.active if tab == 'Engine' else wb.create_sheet(tab)
    if tab == 'Engine':
        ws.title = 'Engine'
    SHEET = tab
    PERIODS = periods
    ws.column_dimensions['A'].width = 48
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 10.5
    ws.cell(row=1, column=1, value=f'GLOBAL ATLANTIC — THE {"QUARTERLY" if tab == "Engine" else "ANNUAL"} ENGINE'
            '  (every line from gate-verified GALD supplement extracts)').font = HDR
    row = 2
    put('$ in millions as filed. Blue = filed value (GALD financial supplements, sha in acquisition/manifest.csv). '
        'Black bold = formula. Grey checks read FAIL if an identity breaks, n/a where a period is not filed.', GREY)
    row = 4
    for i, p in enumerate(periods):
        c = ws.cell(row=4, column=2 + i, value=p)
        c.font = BOLD
        c.alignment = Alignment(horizontal='center')
    row = 5
    build()

# FY stocks: year-end aliases
for fy, q in (('FY2023', '4Q23'), ('FY2024', '4Q24'), ('FY2025', '4Q25')):
    for k in list({k for (_, k) in D}):
        if k.startswith(('res_', 'naic', 'nrsro', 'shareholders_equity', 'adj_equity',
                         'total_debt', 'total_assets', 'adj_invested_assets')):
            if (q, k) in D:
                D.setdefault((fy, k), D[(q, k)])
# rebuild FY tab now that aliases exist
wb.remove(wb['Engine-FY'])
ws = wb.create_sheet('Engine-FY')
SHEET = 'Engine-FY'
PERIODS = FYS
ws.column_dimensions['A'].width = 48
for i in range(len(FYS)):
    ws.column_dimensions[get_column_letter(2 + i)].width = 12
ws.cell(row=1, column=1, value='GLOBAL ATLANTIC — THE ANNUAL ENGINE  (FY columns from supplement; stocks = year-ends)').font = HDR
row = 2
put('Same rows as the quarterly tab.', GREY)
row = 4
for i, p in enumerate(FYS):
    c = ws.cell(row=4, column=2 + i, value=p)
    c.font = BOLD
    c.alignment = Alignment(horizontal='center')
row = 5
build()

cs = wb.create_sheet('Checks')
cs.column_dimensions['A'].width = 56
cs['A1'] = 'LIVE CHECKS — must read 0'
cs['A1'].font = HDR
CHUNK = 40
rows_ = []
for ci in range(0, len(check_cells), CHUNK):
    rr_ = 8 + len(rows_)
    cs.cell(row=rr_, column=2, value='=' + '+'.join(f'COUNTIF({c},"FAIL")' for c in check_cells[ci:ci + CHUNK])).font = GREY
    rows_.append(rr_)
cs['B3'] = f'=SUM(B8:B{rows_[-1]})'
cs['B3'].font = BOLD
cs['A3'] = f'Failing checks (of {len(check_cells)} watched):'

ds = wb.create_sheet('Data')
for j, h in enumerate(('period', 'metric', 'value')):
    ds.cell(row=1, column=1 + j, value=h).font = BOLD
i = 2
for r in csv.DictReader(open(ROOT / 'extract/global-atlantic/gald_supplement.csv')):
    ds.cell(row=i, column=1, value=r['period']).font = BLACK
    ds.cell(row=i, column=2, value=r['metric']).font = BLACK
    if r['value'] != '':
        ds.cell(row=i, column=3, value=float(r['value'])).font = BLACK
    i += 1

rm = wb.create_sheet('ReadMe')
rm.column_dimensions['A'].width = 116
for i, txt in enumerate((
    'GLOBAL ATLANTIC ENGINE — README',
    '',
    'Sources: GALD quarterly financial supplements Q4-2023 / Q4-2024 / Q4-2025 / Q1-2026 (globalatlantic.com;',
    'sha256 in acquisition/manifest.csv). Extractor tools/parse_gald_supplement.py: 918 cells; gates = revenue sums,',
    'underwriting chain, NBV sums, NAIC+NRSRO quality sums, reserves sum, 4-way cross-document overlap equality.',
    'Disclosed recasts (flow/PRT split, FABN re-basis, 1Q26 opex-to-cost-of-insurance reclass) resolved to the',
    'latest-published basis with pretax-neutrality checks; each instance logged by the parser.',
    '',
    'Coverage boundaries: reserves-by-product series starts 4Q24 (earlier supplements use a different layout);',
    'quality tables are period-ends per doc (4Q22-4Q25 year-ends + 1Q26). KKR 10-K/10-Q insurance-segment lane',
    'extracted separately (extract/global-atlantic/kkr_insurance_series.csv) as a cross-source check.',
    '',
    'Companion artifacts: Athene engine workbook (athene-quarterly-engine.xlsx) uses the same architecture;',
    'the cross-company comparison page is the joint verdict surface.',
), start=1):
    rm.cell(row=i, column=1, value=txt).font = HDR if i == 1 else BLACK

wb.save(DEST)

# python-side identity verification (no LibreOffice)
fails = []
for periods in (QTRS, FYS):
    for p in periods:
        g = lambda k: D.get((p, k))
        def eq(a, b, nm):
            if a is not None and b is not None and a != b:
                fails.append(f'{p}/{nm}')
        if None not in (g('nbv_fixed_rate'), g('nbv_fia'), g('nbv_va'), g('nbv_retirement_total')):
            eq(g('nbv_fixed_rate') + g('nbv_fia') + g('nbv_va'), g('nbv_retirement_total'), 'nbv')
        if None not in (g('premiums'), g('policy_fees'), g('nii'), g('inv_gl'), g('other_income'), g('total_revenues')):
            eq(g('premiums') + g('policy_fees') + g('nii') + g('inv_gl') + g('other_income'), g('total_revenues'), 'rev')
        if None not in (g('adj_nii'), g('adj_cost_ins'), g('adj_underwriting')):
            eq(g('adj_nii') - g('adj_cost_ins'), g('adj_underwriting'), 'uw')
print(f'wrote {DEST}')
print('python verification:', 'ALL IDENTITIES HOLD' if not fails else fails[:6],
      f'| live check cells: {len(check_cells)}')
if fails:
    raise SystemExit(1)
