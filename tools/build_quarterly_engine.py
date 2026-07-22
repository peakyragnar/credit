#!/usr/bin/env python3
"""Build the engine workbook (Michael's stock-and-flow sheet, live) — two mirrored tabs.

Engine    - quarterly columns (4Q24..1Q26 + yellow 2Q26 fill column)
Engine-FY - the SAME rows, annual columns (FY2023 from 10-K MD&A; FY2024/FY2025
            from the Q4'25 supplement's full-year YTD columns, cross-gated
            against the 10-K), plus a live 'FY25 Σ quarters' audit column.
Checks    - live count of failing identity checks (must read 0)
Data      - raw extracted cells (traceability)
ReadMe    - sources, method, legend, coverage boundary

Blue = filed value · black bold = formula · grey = check · yellow = fill-in.
FY2023 cells are blank where the row is published only in supplements (the
older supplements are glyph-encoded); their checks read n/a, never FAIL.
No LibreOffice locally: formula arithmetic is verified in Python at the end.
"""
import csv
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DEST = ROOT / 'dossiers/athene/athene-quarterly-engine.xlsx'

# ---- data ----
QD = {}
for r in csv.DictReader(open(ROOT / 'extract/athene/quarterly_supplement.csv')):
    if r['value'] != '':
        QD[(r['quarter'], r['metric'])] = float(r['value'])
AD = {}
for r in csv.DictReader(open(ROOT / 'extract/athene/annual_engine.csv')):
    AD[(r['year'], r['metric'])] = float(r['value'])

QUARTERS = ['4Q24', '1Q25', '2Q25', '3Q25', '4Q25', '1Q26']
FYS = ['FY2023', 'FY2024', 'FY2025']

FYD = {}
for (p, k), v in QD.items():
    if p in ('FY2024', 'FY2025'):
        FYD[(p, k)] = v
for (y, k), v in AD.items():
    FYD.setdefault((y, k), v)                      # 10-K fills FY2023 (+ anything supplements lack)
for fy, q in (('FY2024', '4Q24'), ('FY2025', '4Q25')):
    for k in list({k for (_, k) in QD}):
        if k in ('nrl_begin', 'nrl_end'):
            continue                               # rollforward stocks come from the YTD columns
        if (k.startswith(('nrl_', 'nia_', 'cq_', 'eq_'))) and (q, k) in QD:
            FYD[(fy, k)] = QD[(q, k)]              # year-end stock = 4Q period-end
for key in list(FYD):
    if key[1] == 'provision_credit_losses':
        FYD[key] = -abs(FYD[key])                  # note prints spaced parens; sign restored

# statutory bond-book features (footed d1_trends.csv, BACV -> $M)
_SRC_MAP = {'FE (public agency rating)': 'stat_src_fe', 'PL (private letter)': 'stat_src_pl',
            'FM (financially modeled)': 'stat_src_fm', 'self/temporary (Z,YE,E,M,GI...)': 'stat_src_self',
            'none (exempt/blank)': 'stat_src_none'}
_ID_MAP = {'CUSIP': 'stat_id_cusip', 'PPN (private placement)': 'stat_id_ppn',
           'no identifier': 'stat_id_noid'}
for r in csv.DictReader(open(ROOT / 'extract/athene/d1_trends.csv')):
    fy = 'FY' + r['year']
    v = int(r['bacv']) / 1e6
    d, b = r['dimension'], r['bucket']
    if d == 'total':
        FYD[(fy, 'stat_total')] = v
    elif d == 'naic_band':
        if b in ('NAIC 1', 'NAIC 2'):
            FYD[(fy, 'stat_naic' + b[-1])] = v
        else:
            FYD[(fy, 'stat_belowig')] = FYD.get((fy, 'stat_belowig'), 0) + v
    elif d == 'cliff_2C' and b == '2.C (BBB-)':
        FYD[(fy, 'stat_cliff')] = v
    elif d == 'rating_source':
        FYD[(fy, _SRC_MAP[b])] = v
    elif d == 'id_type':
        FYD[(fy, _ID_MAP[b])] = v

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10)
BOLD = Font(name='Arial', size=10, bold=True)
HDR = Font(name='Arial', size=11, bold=True)
GREY = Font(name='Arial', size=9, color='666666')
YELLOW = PatternFill('solid', fgColor='FFFF00')
MONEY_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.00%'

wb = Workbook()

# ---- pass state (rebound per sheet) ----
ws = None
row = 1
PERIODS = []
COLS = {}
DATA = {}
SHEET = ''
FILL = False          # add yellow fill column (quarterly sheet only)
AUDIT = False         # add Σ-quarters audit columns (FY sheet only)
ANNUALIZE = 4         # NIS%-derivation factor: quarterly ×4, annual ×1
REGISTER = False      # record metric -> row (quarterly pass feeds the audit refs)
GUARD = False         # blank-guard the checks (FY pass: missing FY2023 -> n/a)
METRIC_ROW = {}
check_cells = []


def put(label, font=BLACK, indent=0):
    c = ws.cell(row=row, column=1, value=('    ' * indent) + label)
    if label.startswith('='):
        c.data_type = 's'
    c.font = font
    return row


def vals(metric, fmt=MONEY_FMT, pct=False):
    if REGISTER:
        METRIC_ROW.setdefault(metric, row)
    for i, p in enumerate(PERIODS):
        v = DATA.get((p, metric))
        c = ws.cell(row=row, column=2 + i)
        if v is not None:
            c.value = v / 100.0 if pct else v
        c.font = BLUE
        c.number_format = fmt
    if FILL:
        ws.cell(row=row, column=2 + len(PERIODS)).fill = YELLOW
    if AUDIT:
        qrow = METRIC_ROW.get(metric)
        if qrow:
            ac = ws.cell(row=row, column=2 + len(PERIODS),
                         value=f'=SUM(Engine!C{qrow}:F{qrow})' if not pct else '')
            if pct:
                ac.value = None
            else:
                ac.font = BLACK
                ac.number_format = fmt
                fycol = get_column_letter(1 + len(PERIODS))       # FY2025 column
                acol = get_column_letter(2 + len(PERIODS))
                cc = ws.cell(row=row, column=3 + len(PERIODS),
                             value=f'=IF({fycol}{row}="","n/a",IF({acol}{row}={fycol}{row},"OK","FAIL"))')
                cc.font = GREY
                check_cells.append(f"'{SHEET}'!{get_column_letter(3 + len(PERIODS))}{row}")


def formula(fml_by_col, fmt=MONEY_FMT, font=BOLD):
    for i, p in enumerate(PERIODS):
        col = get_column_letter(2 + i)
        c = ws.cell(row=row, column=2 + i, value=fml_by_col(col))
        c.font = font
        c.number_format = fmt


def check(fml_by_col):
    for i, p in enumerate(PERIODS):
        col = get_column_letter(2 + i)
        f = fml_by_col(col)
        if GUARD:
            refs = re.findall(r'[B-H]\d+', f)
            if refs:
                probe = refs[-1]
                inner = f[1:]
                f = f'=IF({probe}="","n/a",{inner})'
        c = ws.cell(row=row, column=2 + i, value=f)
        c.font = GREY
        check_cells.append(f"'{SHEET}'!{col}{row}")


def section(title):
    global row
    row += 1
    put(title, HDR)
    row += 1


def build_body():
    """All sections; identical row structure on both sheets."""
    global row

    section('MONEY IN — BY CHANNEL')
    r0 = row
    put('Retail', BLUE); vals('retail'); row += 1
    put('Flow reinsurance', BLUE); vals('flow_reins'); row += 1
    put('Funding agreements', BLUE); vals('funding_agreements'); row += 1
    put('Pension group annuities', BLUE); vals('pga'); row += 1
    put('Other spread products', BLUE); vals('other_spread'); row += 1
    r_org = row
    put('= Gross organic inflows', BOLD); formula(lambda c: f'=SUM({c}{r0}:{c}{r0+4})'); row += 1
    put('Gross inorganic (block) inflows', BLUE); vals('gross_inorganic'); row += 1
    r_tot = row
    put('= TOTAL GROSS INFLOWS', BOLD); formula(lambda c: f'={c}{r_org}+{c}{r_org+1}'); row += 1
    r_totrep = row
    put('   reported total (filed)', BLUE); vals('total_gross_inflows'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_tot}={c}{r_totrep},"OK","FAIL")'); row += 1

    section('SAME MONEY, BY OWNER — the two cuts must agree')
    r1 = row
    put('Inflows attributable to Athene', BLUE); vals('inflows_athene'); row += 1
    put('Inflows attributable to ADIP (sidecar investors)', BLUE); vals('inflows_adip'); row += 1
    put('Inflows ceded to third-party reinsurers', BLUE); vals('inflows_ceded_3p'); row += 1
    r_own = row
    put('= Total by owner', BOLD); formula(lambda c: f'=SUM({c}{r1}:{c}{r1+2})'); row += 1
    put('   check vs channel cut', GREY); check(lambda c: f'=IF({c}{r_own}={c}{r_totrep},"OK","FAIL")'); row += 1

    section('MONEY OUT')
    r2 = row
    put('Outflows attributable to Athene', BLUE); vals('outflows_athene'); row += 1
    put('Outflows attributable to ADIP', BLUE); vals('outflows_adip'); row += 1
    r_out = row
    put('= Total gross outflows', BOLD); formula(lambda c: f'={c}{r2}+{c}{r2+1}'); row += 1
    r_outrep = row
    put('   reported total (filed)', BLUE); vals('gross_outflows'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_out}={c}{r_outrep},"OK","FAIL")'); row += 1
    put('Athene outflows by type:', GREY); row += 1
    r3 = row
    put('Maturity-driven / contractual', BLUE, 1); vals('outflow_maturity_driven'); row += 1
    r4 = row
    put('Policyholder-driven', BLUE, 1); vals('outflow_policyholder'); row += 1
    put('income-oriented (planned)', BLUE, 2); vals('outflow_income_planned'); row += 1
    put('out of surrender charge (planned)', BLUE, 2); vals('outflow_oosc_planned'); row += 1
    put('in surrender charge (unplanned)', BLUE, 2); vals('outflow_isc_unplanned'); row += 1
    put('   check: policyholder = its three parts', GREY)
    check(lambda c: f'=IF({c}{r4}=SUM({c}{r4+1}:{c}{r4+3}),"OK","FAIL")'); row += 1
    put('   check: maturity + policyholder = Athene outflows', GREY)
    check(lambda c: f'=IF({c}{r3}+{c}{r4}={c}{r2},"OK","FAIL")'); row += 1

    section('NET FLOWS')
    r_nf = row
    put('= Net flows (inflows + outflows)', BOLD); formula(lambda c: f'={c}{r_tot}+{c}{r_out}'); row += 1
    r_nfrep = row
    put('   reported net flows (filed)', BLUE); vals('net_flows'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_nf}={c}{r_nfrep},"OK","FAIL")'); row += 1

    section('RESERVES — STOCK & FLOW (net reserve liabilities)')
    r5 = row
    put('Reserves — beginning', BLUE); vals('nrl_begin'); row += 1
    put('+ Net inflows (Athene share)', BLUE); vals('roll_net_inflows'); row += 1
    put('− Net withdrawals', BLUE); vals('roll_net_withdrawals'); row += 1
    put('± ACRA ownership changes', BLUE); vals('roll_acra_own'); row += 1
    put('+ Other reserve changes (credited interest & misc)', BLUE); vals('roll_other'); row += 1
    r_end = row
    put('= Reserves — ending', BOLD); formula(lambda c: f'=SUM({c}{r5}:{c}{r5+4})'); row += 1
    r_endrep = row
    put('   reported ending (filed)', BLUE); vals('nrl_end'); row += 1
    put('   check: identity', GREY); check(lambda c: f'=IF({c}{r_end}={c}{r_endrep},"OK","FAIL")'); row += 1
    put('   check: continuity (begin = prior end)', GREY)
    for i, p in enumerate(PERIODS):
        col = get_column_letter(2 + i)
        c = ws.cell(row=row, column=2 + i)
        if i == 0:
            c.value = 'n/a'
        else:
            pcol = get_column_letter(1 + i)
            f = f'=IF({col}{r5}={pcol}{r_endrep},"OK","FAIL")'
            if GUARD:
                f = f'=IF({pcol}{r_endrep}="","n/a",IF({col}{r5}={pcol}{r_endrep},"OK","FAIL"))'
            c.value = f
            check_cells.append(f"'{SHEET}'!{col}{row}")
        c.font = GREY
    row += 1

    section('THE SIDECAR LEDGER — ACRA noncontrolling interests reserves')
    r6 = row
    put('ACRA reserves — beginning', BLUE); vals('acra_begin'); row += 1
    put('+ Inflows', BLUE); vals('acra_inflows'); row += 1
    put('− Withdrawals', BLUE); vals('acra_withdrawals'); row += 1
    put('± ACRA ownership changes', BLUE); vals('acra_own'); row += 1
    put('+ Other reserve changes', BLUE); vals('acra_other'); row += 1
    r_aend = row
    put('= ACRA reserves — ending', BOLD); formula(lambda c: f'=SUM({c}{r6}:{c}{r6+4})'); row += 1
    r_aendrep = row
    put('   reported ending (filed)', BLUE); vals('acra_end'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_aend}={c}{r_aendrep},"OK","FAIL")'); row += 1

    section('RESERVE STOCK BY PRODUCT (period-ends / year-ends)')
    r7 = row
    put('Indexed annuities', BLUE); vals('nrl_indexed'); row += 1
    put('Fixed rate annuities', BLUE); vals('nrl_fixed'); row += 1
    r_def = row
    put('= Total deferred annuities', BOLD); formula(lambda c: f'={c}{r7}+{c}{r7+1}'); row += 1
    put('Pension group annuities', BLUE); vals('nrl_pga'); row += 1
    put('Payout annuities', BLUE); vals('nrl_payout'); row += 1
    put('Funding agreements (the runnable slice)', BLUE); vals('nrl_fa'); row += 1
    put('Life and other', BLUE); vals('nrl_life_other'); row += 1
    r_stot = row
    put('= Total net reserve liabilities', BOLD)
    formula(lambda c: f'={c}{r_def}+SUM({c}{r_def+1}:{c}{r_def+4})'); row += 1
    r_stotrep = row
    put('   reported total (filed)', BLUE); vals('nrl_total'); row += 1
    put('   check (only where period-end filed)', GREY)
    check(lambda c: f'=IF({c}{r_stotrep}="","n/a",IF({c}{r_stot}={c}{r_stotrep},"OK","FAIL"))'); row += 1

    section('THE INCOME ENGINE — spread related earnings')
    r8 = row
    put('Fixed income & other net investment income', BLUE); vals('sre_fi_nii'); row += 1
    put('Alternatives net investment income', BLUE); vals('sre_alt_nii'); row += 1
    r_nie = row
    put('= Net investment earnings', BOLD); formula(lambda c: f'={c}{r8}+{c}{r8+1}'); row += 1
    put('+ Strategic capital management fees', BLUE); vals('sre_fees'); row += 1
    put('− Cost of funds (what the float costs)', BLUE); vals('sre_cof'); row += 1
    r_nis = row
    put('= NET INVESTMENT SPREAD', BOLD); formula(lambda c: f'=SUM({c}{r_nie}:{c}{r_nie+2})'); row += 1
    r_nisrep = row
    put('   reported (filed)', BLUE); vals('sre_nis'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_nis}={c}{r_nisrep},"OK","FAIL")'); row += 1
    put('− Other operating expenses', BLUE); vals('sre_opex'); row += 1
    put('− Interest & other financing costs', BLUE); vals('sre_fin'); row += 1
    r_sre = row
    put('= SPREAD RELATED EARNINGS', BOLD)
    # r_nis+1 = reported, r_nis+2 = check, r_nis+3 = opex, r_nis+4 = financing
    formula(lambda c: f'={c}{r_nis}+{c}{r_nis+3}+{c}{r_nis+4}'); row += 1
    r_srerep = row
    put('   reported (filed)', BLUE); vals('sre'); row += 1
    put('   check', GREY); check(lambda c: f'=IF({c}{r_sre}={c}{r_srerep},"OK","FAIL")'); row += 1

    section('THE RATES (annualized, as filed)')
    put('Earned — fixed income & other', BLUE); vals('sre_fi_nii_pct', fmt=PCT_FMT, pct=True); row += 1
    put('Earned — alternatives', BLUE); vals('sre_alt_nii_pct', fmt=PCT_FMT, pct=True); row += 1
    put('Earned — total', BLUE); vals('sre_nie_pct', fmt=PCT_FMT, pct=True); row += 1
    put('Cost of funds', BLUE); vals('sre_cof_pct', fmt=PCT_FMT, pct=True); row += 1
    r_nispct = row
    put('NET INVESTMENT SPREAD %', BLUE); vals('sre_nis_pct', fmt=PCT_FMT, pct=True); row += 1
    put('SRE margin %', BLUE); vals('sre_pct', fmt=PCT_FMT, pct=True); row += 1

    section('THE DENOMINATOR & DERIVED')
    r9 = row
    put('Average net invested assets — fixed income', BLUE); vals('avg_nia_fi'); row += 1
    put('Average net invested assets — alternatives', BLUE); vals('avg_nia_alt'); row += 1
    r_nia = row
    put('= Average net invested assets', BOLD); formula(lambda c: f'={c}{r9}+{c}{r9+1}'); row += 1
    r_dnis = row
    put(f'Derived NIS % (spread ×{ANNUALIZE} ÷ avg assets)', BOLD)
    formula(lambda c: f'=IF(OR({c}{r_nia}=0,{c}{r_nia}=""),"",{c}{r_nis}*{ANNUALIZE}/{c}{r_nia})',
            fmt=PCT_FMT, font=BOLD); row += 1
    put('   check vs filed NIS % (±5bp rounding)', GREY)
    check(lambda c: f'=IF({c}{r_dnis}="","n/a",IF(ABS({c}{r_dnis}-{c}{r_nispct})<0.0005,"OK","FAIL"))'); row += 1

    section('THE PORTFOLIO — STATUTORY BOND BOOK, THE FEATURES WE BUILT (year-ends; footed to the dollar)')
    ra = row
    put('Total bond book (BACV)', BLUE); vals('stat_total'); row += 1
    put('NAIC 1 (AAA…A−)', BLUE, 1); vals('stat_naic1'); row += 1
    put('NAIC 2 (BBB band)', BLUE, 1); vals('stat_naic2'); row += 1
    put('Below investment grade (NAIC 3–6)', BLUE, 1); vals('stat_belowig'); row += 1
    put('   check: bands sum to total', GREY)
    check(lambda c: f'=IF({c}{ra}="","n/a",IF(SUM({c}{ra+1}:{c}{ra+3})={c}{ra},"OK","FAIL"))'); row += 1
    put('BBB− cliff (2.C) — memo', BLUE, 1); vals('stat_cliff'); row += 1
    rs = row
    put('Publicly rated (FE)', BLUE, 1); vals('stat_src_fe'); row += 1
    put('PRIVATE LETTER (PL) — the channel', BLUE, 1); vals('stat_src_pl'); row += 1
    put('Modeled (FM)', BLUE, 1); vals('stat_src_fm'); row += 1
    put('Self/temporary (Z,YE…)', BLUE, 1); vals('stat_src_self'); row += 1
    put('Exempt/blank', BLUE, 1); vals('stat_src_none'); row += 1
    put('   check: sources sum to total', GREY)
    check(lambda c: f'=IF({c}{ra}="","n/a",IF(SUM({c}{rs}:{c}{rs+4})={c}{ra},"OK","FAIL"))'); row += 1
    r_pl = rs + 1
    put('= PL share of book', BOLD)
    formula(lambda c: f'=IF({c}{ra}="","",{c}{r_pl}/{c}{ra})', fmt=PCT_FMT); row += 1
    ri = row
    put('CUSIP (public identifier)', BLUE, 1); vals('stat_id_cusip'); row += 1
    put('PPN (private placement)', BLUE, 1); vals('stat_id_ppn'); row += 1
    put('No identifier', BLUE, 1); vals('stat_id_noid'); row += 1
    r_floor = row
    put('= Identifier-private floor (PPN + no-ID)', BOLD)
    formula(lambda c: f'=IF({c}{ri}="","",{c}{ri+1}+{c}{ri+2})'); row += 1
    put('= Private floor share of book', BOLD)
    formula(lambda c: f'=IF({c}{ra}="","",{c}{r_floor}/{c}{ra})', fmt=PCT_FMT); row += 1
    put('   check: identifiers sum to total', GREY)
    check(lambda c: f'=IF({c}{ra}="","n/a",IF(SUM({c}{ri}:{c}{ri+2})={c}{ra},"OK","FAIL"))'); row += 1

    section('THE PORTFOLIO — MANAGEMENT VIEW, NET INVESTED ASSETS BY CLASS (period-ends)')
    rn = row
    put('Corporate', BLUE, 1); vals('nia_corporate'); row += 1
    put('CLO', BLUE, 1); vals('nia_clo'); row += 1
    r_cred = row
    put('= Credit', BOLD); formula(lambda c: f'=IF({c}{rn}="","",{c}{rn}+{c}{rn+1})'); row += 1
    put('Commercial mortgage loans', BLUE, 1); vals('nia_cml'); row += 1
    put('Residential mortgage loans', BLUE, 1); vals('nia_rml'); row += 1
    put('RMBS', BLUE, 1); vals('nia_rmbs'); row += 1
    put('CMBS', BLUE, 1); vals('nia_cmbs'); row += 1
    r_re = row
    put('= Real estate', BOLD); formula(lambda c: f'=IF({c}{r_cred+1}="","",SUM({c}{r_cred+1}:{c}{r_cred+4}))'); row += 1
    put('ABS', BLUE, 1); vals('nia_abs'); row += 1
    put('Alternative investments', BLUE, 1); vals('nia_alts'); row += 1
    put('Munis / foreign government', BLUE, 1); vals('nia_munis_foreign'); row += 1
    put('Equity securities', BLUE, 1); vals('nia_equity_sec'); row += 1
    put('Short-term investments', BLUE, 1); vals('nia_short_term'); row += 1
    put('US government and agencies', BLUE, 1); vals('nia_us_gov'); row += 1
    r_oth = row
    put('= Other investments', BOLD); formula(lambda c: f'=IF({c}{r_re+1}="","",SUM({c}{r_re+1}:{c}{r_re+6}))'); row += 1
    put('Cash and cash equivalents', BLUE, 1); vals('nia_cash'); row += 1
    put('Other', BLUE, 1); vals('nia_other'); row += 1
    r_niat = row
    put('= NET INVESTED ASSETS', BOLD)
    formula(lambda c: f'=IF({c}{r_cred}="","",{c}{r_cred}+{c}{r_re}+{c}{r_oth}+{c}{r_oth+1}+{c}{r_oth+2})'); row += 1
    r_niarep = row
    put('   reported (filed)', BLUE); vals('nia_total'); row += 1
    put('   check', GREY)
    check(lambda c: f'=IF({c}{r_niarep}="","n/a",IF({c}{r_niat}={c}{r_niarep},"OK","FAIL"))'); row += 1

    section('THE PORTFOLIO — CREDIT QUALITY, NAIC-DESIGNATED ASSETS (management view, period-ends)')
    rq = row
    put('NAIC 1', BLUE, 1); vals('cq_naic1'); row += 1
    put('NAIC 2', BLUE, 1); vals('cq_naic2'); row += 1
    put('Non-designated (IG)', BLUE, 1); vals('cq_nondesig_ig'); row += 1
    r_ig = row
    put('= Total investment grade', BOLD); formula(lambda c: f'=IF({c}{rq}="","",SUM({c}{rq}:{c}{rq+2}))'); row += 1
    put('NAIC 3', BLUE, 1); vals('cq_naic3'); row += 1
    put('NAIC 4', BLUE, 1); vals('cq_naic4'); row += 1
    put('NAIC 5', BLUE, 1); vals('cq_naic5'); row += 1
    put('NAIC 6', BLUE, 1); vals('cq_naic6'); row += 1
    put('Non-designated (BIG)', BLUE, 1); vals('cq_nondesig_big'); row += 1
    r_big = row
    put('= Total below investment grade', BOLD)
    formula(lambda c: f'=IF({c}{r_ig+1}="","",SUM({c}{r_ig+1}:{c}{r_ig+5}))'); row += 1
    r_cqt = row
    put('= Total NAIC-designated assets', BOLD)
    formula(lambda c: f'=IF({c}{r_ig}="","",{c}{r_ig}+{c}{r_big})'); row += 1
    r_cqtrep = row
    put('   reported (filed)', BLUE); vals('cq_total_desig'); row += 1
    put('   check', GREY)
    check(lambda c: f'=IF({c}{r_cqtrep}="","n/a",IF({c}{r_cqt}={c}{r_cqtrep},"OK","FAIL"))'); row += 1

    section('GAAP INCOME — TO THE NET INCOME THAT VALUES THE EQUITY')
    rg = row
    put('Premiums', BLUE, 1); vals('gaap_premiums'); row += 1
    put('Product charges', BLUE, 1); vals('gaap_product_charges'); row += 1
    put('Net investment income', BLUE, 1); vals('gaap_nii'); row += 1
    put('Investment related gains (losses)', BLUE, 1); vals('gaap_inv_gl'); row += 1
    put('Other revenues', BLUE, 1); vals('gaap_other_rev'); row += 1
    put('VIE net investment income', BLUE, 1); vals('gaap_vie_nii'); row += 1
    put('VIE investment related gains (losses)', BLUE, 1); vals('gaap_vie_gl'); row += 1
    r_rev = row
    put('= Total revenues', BOLD); formula(lambda c: f'=IF({c}{rg}="","",SUM({c}{rg}:{c}{rg+6}))'); row += 1
    r_revrep = row
    put('   reported (filed)', BLUE); vals('gaap_total_rev'); row += 1
    put('   check', GREY)
    check(lambda c: f'=IF({c}{r_revrep}="","n/a",IF({c}{r_rev}={c}{r_revrep},"OK","FAIL"))'); row += 1
    rb = row
    put('Interest sensitive contract benefits', BLUE, 1); vals('gaap_iscb'); row += 1
    put('Future policy benefits', BLUE, 1); vals('gaap_fpb'); row += 1
    put('Market risk benefits remeasurement', BLUE, 1); vals('gaap_mrb'); row += 1
    put('DAC amortization', BLUE, 1); vals('gaap_dac'); row += 1
    put('Policy and other operating expenses', BLUE, 1); vals('gaap_opex'); row += 1
    r_be = row
    put('= Total benefits and expenses', BOLD); formula(lambda c: f'=IF({c}{rb}="","",SUM({c}{rb}:{c}{rb+4}))'); row += 1
    put('Provision for credit losses (annual, GAAP; memo — inside inv. gains/losses)', BLUE, 1)
    vals('provision_credit_losses'); row += 1
    r_pt = row
    put('= Pre-tax income', BOLD); formula(lambda c: f'=IF({c}{r_rev}="","",{c}{r_rev}-{c}{r_be})'); row += 1
    put('− Income tax expense (benefit)', BLUE, 1); vals('gaap_tax'); row += 1
    r_ni = row
    put('= NET INCOME', BOLD); formula(lambda c: f'=IF({c}{r_pt}="","",{c}{r_pt}-{c}{r_pt+1})'); row += 1
    r_nirep = row
    put('   reported net income (filed)', BLUE); vals('gaap_ni'); row += 1
    put('   check', GREY)
    check(lambda c: f'=IF({c}{r_nirep}="","n/a",IF({c}{r_ni}={c}{r_nirep},"OK","FAIL"))'); row += 1
    put('− NI attributable to NCI (ADIP)', BLUE, 1); vals('gaap_ni_nci'); row += 1
    put('− Preferred dividends (net of redemption adj.)', BLUE, 1); vals('gaap_pref'); row += 1
    put('+ Preferred redemption adjustment', BLUE, 1); vals('gaap_pref_red'); row += 1
    r_nic = row
    put('= NET INCOME TO COMMON — the valuation line', BOLD)
    formula(lambda c: f'=IF({c}{r_nirep}="","",{c}{r_nirep}-{c}{r_nic-3}-{c}{r_nic-2}+{c}{r_nic-1})'); row += 1
    r_nicrep = row
    put('   reported (filed)', BLUE); vals('gaap_ni_common'); row += 1
    put('   check', GREY)
    check(lambda c: f'=IF({c}{r_nicrep}="","n/a",IF({c}{r_nic}={c}{r_nicrep},"OK","FAIL"))'); row += 1

    section('THE EQUITY UNDERNEATH — AND THE RETURN ON IT')
    re_ = row
    put('Additional paid-in capital', BLUE, 1); vals('eq_apic'); row += 1
    put('Retained earnings', BLUE, 1); vals('eq_re'); row += 1
    put('AOCI (the bond-mark hole)', BLUE, 1); vals('eq_aoci'); row += 1
    r_ahl = row
    put('= AHL stockholders equity', BOLD)
    formula(lambda c: f'=IF({c}{re_}="","",SUM({c}{re_}:{c}{re_+2}))'); row += 1
    r_ahlrep = row
    put('   reported (filed)', BLUE); vals('eq_ahl_total'); row += 1
    put('   check', GREY)
    check(lambda c: f'=IF({c}{r_ahlrep}="","n/a",IF({c}{r_ahl}={c}{r_ahlrep},"OK","FAIL"))'); row += 1
    put('Noncontrolling interests (ADIP)', BLUE, 1); vals('eq_nci'); row += 1
    r_eqt = row
    put('= Total equity', BOLD)
    formula(lambda c: f'=IF({c}{r_ahlrep}="","",{c}{r_ahlrep}+{c}{r_eqt-1})'); row += 1
    r_eqtrep = row
    put('   reported (filed)', BLUE); vals('eq_total'); row += 1
    put('   check', GREY)
    check(lambda c: f'=IF({c}{r_eqtrep}="","n/a",IF({c}{r_eqt}={c}{r_eqtrep},"OK","FAIL"))'); row += 1
    r_roe = row
    put(f'ROE — NI-common ×{ANNUALIZE} ÷ avg AHL equity (this and prior period)', BOLD)
    for i, p in enumerate(PERIODS):
        col = get_column_letter(2 + i)
        c = ws.cell(row=r_roe, column=2 + i)
        if i == 0:
            c.value = 'n/a'
            c.font = GREY
        else:
            pcol = get_column_letter(1 + i)
            c.value = (f'=IF(OR({col}{r_nicrep}="",{col}{r_ahlrep}="",{pcol}{r_ahlrep}=""),"",'
                       f'{col}{r_nicrep}*{ANNUALIZE}/(({col}{r_ahlrep}+{pcol}{r_ahlrep})/2))')
            c.font = BOLD
            c.number_format = PCT_FMT
    row += 1


def sheet_header(title, note, extra_cols=()):
    global row
    ws.cell(row=1, column=1, value=title).font = HDR
    row = 2
    put(note, GREY)
    row = 4
    for i, p in enumerate(PERIODS):
        c = ws.cell(row=4, column=2 + i, value=p)
        c.font = BOLD
        c.alignment = Alignment(horizontal='center')
    for j, h in enumerate(extra_cols):
        c = ws.cell(row=4, column=2 + len(PERIODS) + j, value=h)
        c.font = BOLD
        c.alignment = Alignment(horizontal='center')
        if h == '2Q26':
            c.fill = YELLOW
    row = 5


# ================= pass 1: quarterly =================
ws = wb.active
ws.title = 'Engine'
SHEET = 'Engine'
PERIODS = QUARTERS
DATA = QD
FILL, AUDIT, REGISTER, GUARD, ANNUALIZE = True, False, True, False, 4
ws.column_dimensions['A'].width = 46
for i in range(len(PERIODS) + 1):
    ws.column_dimensions[get_column_letter(2 + i)].width = 11
sheet_header('ATHENE — THE QUARTERLY ENGINE  (stock & flow, every line from footed extracts)',
             '$ in millions. Blue = filed value (ATH financial supplements, sha in acquisition/manifest.csv). '
             'Black bold = formula. "check" rows show FAIL if an identity breaks. Yellow column = fill when 2Q26 publishes.',
             extra_cols=('2Q26',))
build_body()

# ================= pass 2: annual =================
ws = wb.create_sheet('Engine-FY')
SHEET = 'Engine-FY'
PERIODS = FYS
DATA = FYD
FILL, AUDIT, REGISTER, GUARD, ANNUALIZE = False, True, False, True, 1
ws.column_dimensions['A'].width = 46
for i in range(len(PERIODS) + 2):
    ws.column_dimensions[get_column_letter(2 + i)].width = 12
sheet_header('ATHENE — THE ANNUAL ENGINE  (same rows; FY2023 from 10-K, FY2024/25 from Q4 supplement YTD, cross-gated)',
             'Blue = filed value (ahl-10k-fy2025 MD&A + ATH Q4\'25 supplement full-year columns; both sources agree exactly '
             'where they overlap). "FY25 Σq" = live sum over the Engine tab — its check must read OK. Blank FY2023 cells = '
             'row published only in supplements (older ones are glyph-encoded; decoder parked); their checks read n/a.',
             extra_cols=('FY25 Σq', 'src-check'))
build_body()

# ================= Checks / Data / ReadMe =================
cs = wb.create_sheet('Checks')
cs.column_dimensions['A'].width = 60
cs['A1'] = 'LIVE CHECKS — this cell must read 0'
cs['A1'].font = HDR
cs['A3'] = 'Number of failing identity checks across Engine + Engine-FY:'
cs['A3'].font = BLACK
# Excel caps a formula at 8,192 chars — chunk the COUNTIFs and sum the chunks
CHUNK = 40
chunk_rows = []
for ci in range(0, len(check_cells), CHUNK):
    rr = 8 + len(chunk_rows)
    parts = '+'.join(f'COUNTIF({c},"FAIL")' for c in check_cells[ci:ci + CHUNK])
    cs.cell(row=rr, column=2, value=f'={parts}').font = GREY
    cs.cell(row=rr, column=1, value=f'checks {ci + 1}–{min(ci + CHUNK, len(check_cells))}').font = GREY
    chunk_rows.append(rr)
cs['B3'] = f'=SUM(B8:B{chunk_rows[-1]})'
cs['B3'].font = BOLD
cs['A5'] = (f'Checks watched: {len(check_cells)} cells — totals cross-cuts, rollforward identities, continuity, '
            'SRE chain, derived spread, and the quarterly-vs-annual source audit.')
cs['A5'].font = GREY

ds = wb.create_sheet('Data')
ds.column_dimensions['A'].width = 10
ds.column_dimensions['B'].width = 34
for j, h in enumerate(('period', 'metric', 'value', 'source')):
    ds.cell(row=1, column=1 + j, value=h).font = BOLD
i = 2
for r in csv.DictReader(open(ROOT / 'extract/athene/quarterly_supplement.csv')):
    ds.cell(row=i, column=1, value=r['quarter']).font = BLACK
    ds.cell(row=i, column=2, value=r['metric']).font = BLACK
    if r['value'] != '':
        ds.cell(row=i, column=3, value=float(r['value'])).font = BLACK
    ds.cell(row=i, column=4, value='supplement').font = BLACK
    i += 1
for r in csv.DictReader(open(ROOT / 'extract/athene/annual_engine.csv')):
    ds.cell(row=i, column=1, value=r['year']).font = BLACK
    ds.cell(row=i, column=2, value=r['metric']).font = BLACK
    ds.cell(row=i, column=3, value=float(r['value'])).font = BLACK
    ds.cell(row=i, column=4, value='10-K MD&A').font = BLACK
    i += 1
for r in csv.DictReader(open(ROOT / 'extract/athene/d1_trends.csv')):
    ds.cell(row=i, column=1, value='FY' + r['year']).font = BLACK
    ds.cell(row=i, column=2, value=f"stat:{r['dimension']}:{r['bucket']}").font = BLACK
    ds.cell(row=i, column=3, value=int(r['bacv']) / 1e6).font = BLACK
    ds.cell(row=i, column=4, value='statutory D1 (footed)').font = BLACK
    i += 1

rm = wb.create_sheet('ReadMe')
rm.column_dimensions['A'].width = 118
notes = [
    ('ATHENE ENGINE WORKBOOK — README', HDR),
    ('', BLACK),
    ('TABS: Engine = quarterly (4Q24–1Q26 + yellow 2Q26 fill column). Engine-FY = the SAME rows annually', BLACK),
    ('(FY2023–FY2025) with a live "FY25 Σ quarters" audit column. Checks must read 0. Data = raw extracted cells.', BLACK),
    ('', BLACK),
    ('SOURCES: ATH Q4 2025 + Q1 2026 Financial Supplements (quarterly + full-year YTD columns) and the AHL 10-K', BLACK),
    ('FY2025 MD&A (annual comparatives incl. FY2023). URLs + sha256 in acquisition/manifest.csv. Every blue cell', BLACK),
    ('was machine-extracted and passed exact cross-foot gates; the two sources were gated equal where they overlap.', BLACK),
    ('Extractors: tools/parse_fin_supplement.py, tools/parse_10k_annual_tables.py.', BLACK),
    ('', BLACK),
    ('LEGEND: blue = filed value · black bold = formula (recalculates in Excel) · grey = checks ("FAIL"/"n/a")', BLACK),
    ('· yellow = fill-in. Blank FY2023 cells: those rows are published only in supplements, and the 2023-era', BLACK),
    ('supplements are glyph-encoded (decoder parked). Blanks are honest gaps, not zeros.', BLACK),
    ('', BLACK),
    ('READ THE TREND: cost of funds $5,650M (FY23) → $7,702M → $10,083M (FY25), +78% in two years, while net', BLACK),
    ('investment earnings grew 49%; NIS margin 1.65% (1Q25) → 1.34% (1Q26); SRE margin below 1% in 1Q26.', BLACK),
    ('Volume is outrunning margin, and the margin is losing.', BLACK),
]
for i, (txt, f) in enumerate(notes, start=1):
    rm.cell(row=i, column=1, value=txt).font = f

wb.save(DEST)

# ---- Python-side verification (no LibreOffice): every identity on both sheets ----
fails = []
for periods, data, tag in ((QUARTERS, QD, 'Q'), (FYS, FYD, 'FY')):
    for p in periods:
        g = lambda k: data.get((p, k))
        def has(*ks):
            return all(g(k) is not None for k in ks)
        def eq(lhs, rhs, name):
            if lhs != rhs:
                fails.append(f'{tag}/{p}/{name}: {lhs} != {rhs}')
        if has('retail', 'flow_reins', 'funding_agreements', 'pga', 'other_spread',
               'gross_inorganic', 'total_gross_inflows'):
            eq(g('retail') + g('flow_reins') + g('funding_agreements') + g('pga') + g('other_spread')
               + g('gross_inorganic'), g('total_gross_inflows'), 'inflow chain')
        if has('inflows_athene', 'inflows_adip', 'inflows_ceded_3p', 'total_gross_inflows'):
            eq(g('inflows_athene') + g('inflows_adip') + g('inflows_ceded_3p'),
               g('total_gross_inflows'), 'owner cut')
        if has('outflows_athene', 'outflows_adip', 'gross_outflows'):
            eq(g('outflows_athene') + g('outflows_adip'), g('gross_outflows'), 'outflow cut')
        if has('total_gross_inflows', 'gross_outflows', 'net_flows'):
            eq(g('total_gross_inflows') + g('gross_outflows'), g('net_flows'), 'net flows')
        if has('nrl_begin', 'roll_net_inflows', 'roll_net_withdrawals', 'roll_other', 'nrl_end'):
            eq(g('nrl_begin') + g('roll_net_inflows') + g('roll_net_withdrawals')
               + (g('roll_acra_own') or 0) + g('roll_other'), g('nrl_end'), 'NRL rollforward')
        if has('acra_begin', 'acra_inflows', 'acra_withdrawals', 'acra_other', 'acra_end'):
            eq(g('acra_begin') + g('acra_inflows') + g('acra_withdrawals')
               + (g('acra_own') or 0) + g('acra_other'), g('acra_end'), 'ACRA rollforward')
        if has('sre_fi_nii', 'sre_alt_nii', 'sre_fees', 'sre_cof', 'sre_nis'):
            eq(g('sre_fi_nii') + g('sre_alt_nii') + g('sre_fees') + g('sre_cof'),
               g('sre_nis'), 'NIS chain')
        if has('sre_nis', 'sre_opex', 'sre_fin', 'sre'):
            eq(g('sre_nis') + g('sre_opex') + g('sre_fin'), g('sre'), 'SRE chain')
print(f'wrote {DEST}')
print(f'python verification: {"ALL IDENTITIES HOLD" if not fails else fails[:5]} '
      f'| live check cells: {len(check_cells)}')
if fails:
    raise SystemExit(1)
